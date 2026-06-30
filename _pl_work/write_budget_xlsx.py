from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
SEED = BASE / "budget_seed.json"
OUT = Path(r"C:\Users\PauloMendonça\OneDrive - Redefrete\Área de Trabalho\Balanço\DashBoard_P&L\Budget.xlsx")

NAVY = "061B49"
BLUE = "1763E8"
HEADER = "F3F6FB"
LINE = "D9E2F1"
MUTED = "5B6B8A"


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor=HEADER)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = Font(bold=True, color=MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_sheet(wb, title, headers, rows, widths, currency_cols=None):
    currency_cols = set(currency_cols or [])
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws)
    set_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in currency_cols:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                item.number_format = '"R$" #,##0.00;-"R$" #,##0.00'
    return ws


def main():
    data = json.loads(SEED.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Budget automático - Média dos últimos 4 meses"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:H1")
    rows = [
        ("Data de geração", data["meta"]["generatedAt"]),
        ("Fonte", data["meta"]["source"]),
        ("Meses base", ", ".join(data["meta"]["basePeriodLabels"])),
        ("Meses projetados", ", ".join(data["meta"]["projectPeriodLabels"])),
        ("Combinações base", data["meta"]["baseRows"]),
        ("Linhas de budget", data["meta"]["budgetRows"]),
        ("Linhas de histórico", data["meta"]["historyRows"]),
        ("Cenário", "Budget Média 4M"),
        ("Observação", "Primeira versão sem ajuste de sazonalidade/campanhas."),
    ]
    for row_idx, row in enumerate(rows, start=3):
        ws.cell(row_idx, 1, row[0])
        ws.cell(row_idx, 2, row[1])
    style_header(ws, 3)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A3"

    budget_headers = [
        "Cenário",
        "Ano",
        "Mês",
        "Período",
        "Cliente",
        "Projeto",
        "Unidade",
        "Expt",
        "Tipo",
        "Frota",
        "Conta DRE",
        "Categoria",
        "Tipo Custo",
        "Valor Budget",
        "Observação",
    ]
    budget_rows = [
        [
            row["scenario"],
            row["year"],
            row["month"],
            row["period"],
            row["client"],
            row["project"],
            row["hub"],
            row["expt"],
            row["vehicleType"],
            row["fleetType"],
            row["account"],
            row["category"],
            row["costType"],
            row["budgetValue"],
            row["note"],
        ]
        for row in data["budgetRows"]
    ]
    write_sheet(
        wb,
        "Budget",
        budget_headers,
        budget_rows,
        [18, 10, 9, 12, 18, 22, 20, 22, 14, 12, 30, 34, 14, 16, 34],
        currency_cols=[14],
    )

    base_headers = [
        "Cliente",
        "Projeto",
        "Unidade",
        "Expt",
        "Tipo",
        "Frota",
        "Conta DRE",
        "Categoria",
        "Tipo Custo",
        *data["meta"]["basePeriodLabels"],
        "Média 4M",
        "Meses c/ Movimento",
    ]
    base_rows = [
        [
            row["client"],
            row["project"],
            row["hub"],
            row["expt"],
            row["vehicleType"],
            row["fleetType"],
            row["account"],
            row["category"],
            row["costType"],
            *[row[period] for period in data["meta"]["basePeriods"]],
            row["average"],
            row["activeMonths"],
        ]
        for row in data["baseRows"]
    ]
    write_sheet(
        wb,
        "Base_Media_4M",
        base_headers,
        base_rows,
        [18, 22, 20, 22, 14, 12, 30, 34, 14, 14, 14, 14, 14, 15, 18],
        currency_cols=[10, 11, 12, 13, 14],
    )

    hist_headers = [
        "Período",
        "Cliente",
        "Projeto",
        "Unidade",
        "Expt",
        "Tipo",
        "Frota",
        "Conta DRE",
        "Categoria",
        "Tipo Custo",
        "Valor Realizado",
    ]
    hist_rows = [
        [
            row["period"],
            row["client"],
            row["project"],
            row["hub"],
            row["expt"],
            row["vehicleType"],
            row["fleetType"],
            row["account"],
            row["category"],
            row["costType"],
            row["actualValue"],
        ]
        for row in data["historyRows"]
    ]
    write_sheet(
        wb,
        "Historico_Realizado",
        hist_headers,
        hist_rows,
        [12, 18, 22, 20, 22, 14, 12, 30, 34, 14, 16],
        currency_cols=[11],
    )

    write_sheet(
        wb,
        "Campanhas",
        ["Ano", "Mês", "Cliente", "Projeto", "Unidade", "Campanha", "Fator Sazonal", "Observação"],
        [[2026, "", "", "", "", "", "", "Preencher depois para ajustar o budget por campanha/sazonalidade."]],
        [10, 10, 18, 22, 20, 26, 16, 54],
    )
    wb["Campanhas"]["G2"].number_format = "0.0%"

    write_sheet(
        wb,
        "Premissas",
        ["Indicador", "Dimensão", "Valor / Regra", "Vigência Inicial", "Vigência Final", "Observação"],
        [[
            "Base do Budget",
            "Geral",
            "Média dos últimos 4 meses realizados",
            data["meta"]["projectPeriodLabels"][0],
            data["meta"]["projectPeriodLabels"][-1],
            "Gerado automaticamente; revisar sazonalidade e eventos pontuais.",
        ]],
        [24, 20, 44, 16, 16, 54],
    )

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    check = load_workbook(OUT, read_only=True, data_only=True)
    print(OUT)
    for name in check.sheetnames:
        ws_check = check[name]
        print(name, ws_check.max_row, ws_check.max_column)


if __name__ == "__main__":
    main()
