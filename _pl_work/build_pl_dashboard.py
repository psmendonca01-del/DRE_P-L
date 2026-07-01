import json
import math
import os
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import pl_database


BASE = Path(__file__).resolve().parent
SOURCE_FILE = Path(r"C:\Users\PauloMendonça\OneDrive - Redefrete\Área de Trabalho\Balanço\DashBoard_P&L\P&L.xlsx")
BUDGET_FILE = Path(r"C:\Users\PauloMendonça\OneDrive - Redefrete\Área de Trabalho\Balanço\DashBoard_P&L\Budget.xlsx")
WORKBOOK = BASE / "PL.xlsx"
DATA_OUT = BASE / "pl_data.json"
LEDGER_OUT = BASE / "pl_ledger.json"
HTML_OUT = BASE / "dashboard_pl.html"

MONTHS = {
    "janeiro": "01",
    "fevereiro": "02",
    "março": "03",
    "marco": "03",
    "abril": "04",
    "maio": "05",
    "junho": "06",
    "julho": "07",
    "agosto": "08",
    "setembro": "09",
    "outubro": "10",
    "novembro": "11",
    "dezembro": "12",
}

RATEIO_DEPARTMENTS = {
    "ADMINISTRACAO",
    "FINANCEIRO",
    "ARACOIABA DA SERRA",
    "OPERACOES",
    "RECURSOS HUMANOS",
    "CAPTACAO",
    "T.I",
    "CONTROLADORIA",
    "JURIDICO",
    "MONITORAMENTO",
    "FACILITIES",
    "MARKETING",
    "CONTABILIDADE",
    "SAO PAULO",
}


def safe_float(value):
    if value in (None, ""):
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace(".", "").replace(",", ".") if "," in value else value
        return float(value)
    except Exception:
        return 0.0


def period_from_parts(year, month_name):
    if not year or not month_name:
        return None
    year = str(year).split(".")[0]
    month_key = unicodedata.normalize("NFD", str(month_name).strip().lower())
    month_key = "".join(ch for ch in month_key if unicodedata.category(ch) != "Mn")
    month = MONTHS.get(month_key)
    return f"{year}-{month}" if month else None


def clean(value, fallback="N/D"):
    text = str(value or "").strip()
    return text if text else fallback


def norm(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return " ".join(text.upper().strip().split())


def header_index(ws):
    headers = [norm(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {header: index for index, header in enumerate(headers) if header}


def get_by_header(row, idx, *names, default=None):
    for name in names:
        index = idx.get(norm(name))
        if index is not None and index < len(row):
            return row[index]
    return default


def code_like(value):
    text = norm(value)
    return bool(text) and any(ch.isdigit() for ch in text) and len(text) <= 6 and text.replace(" ", "").isalnum()


def canonical_location(value):
    text = clean(value, "")
    key = norm(text).replace("-", " ")
    key = " ".join(key.split())
    if key in {"ARENA BARUERI", "BARUERI ARENA"}:
        return "Arena Barueri"
    return text


def unit_value(item):
    hub = canonical_location(item.get("hub") or "")
    source_hub = canonical_location(item.get("sourceHub") or "")
    department = canonical_location(item.get("department") or "")
    if hub and not code_like(hub):
        return hub
    if department and department != "N/D" and not code_like(department):
        return department
    if source_hub and not code_like(source_hub):
        return source_hub
    fallback = hub or source_hub or department
    return "Unidade não mapeada" if code_like(fallback) else fallback


def preferred_label(values):
    labels = [clean(value, "") for value in values if clean(value, "")]
    if not labels:
        return ""
    for label in labels:
        if not label.isupper():
            return label
    return labels[0]


def unique_labels(values):
    grouped = defaultdict(list)
    for value in values:
        if value:
            key = "".join(ch for ch in norm(value) if ch.isalnum())
            grouped[key].append(value)
    return sorted(preferred_label(items) for items in grouped.values() if preferred_label(items))


def add_bucket(store, key, values):
    bucket = store[key]
    for name, value in values.items():
        bucket[name] += value


def allocation_key(item):
    return (
        item["client"],
        item["project"],
        item["hub"],
        item.get("expt", "N/D"),
        item.get("sourceHub", item["hub"]),
        item.get("city", "N/D"),
        item.get("department", "N/D"),
        item.get("vehicleType", "N/D"),
        item.get("fleetType", "N/D"),
        item.get("fleetOwner", "N/D"),
    )


def scoped_allocation_keys(period, client, project, hub, expt="Todos"):
    return (
        (period, client, project, hub, expt),
        (period, client, project, hub, "Todos"),
        (period, client, project, hub),
        (period, client, project, "Todos"),
        (period, client, "Todos", "Todos"),
        (period, "Todos", "Todos", "Todos"),
    )


def scoped_rateio_base(period, client, project, hub, expt, scoped_basis, global_basis, is_fleet_utility=False):
    def fleet_only(base):
        return {
            alloc_key: alloc_value
            for alloc_key, alloc_value in base.items()
            if alloc_key[7] == "Frota" and alloc_key[8] == "Propria"
        }

    for scoped_key in scoped_allocation_keys(period, client, project, hub, expt):
        base = scoped_basis.get(scoped_key, {})
        if not base:
            continue
        if is_fleet_utility:
            filtered = fleet_only(base)
            if filtered:
                return filtered
            continue
        return base

    if is_fleet_utility:
        filtered = fleet_only(global_basis)
        if filtered:
            return filtered
    return global_basis


def classify_finance(account, category):
    account = norm(clean(account, "")).lower()
    category = norm(clean(category, "")).lower()
    if "rendimentos de aplicacoes" in account or "rendimentos de aplicacoes" in category:
        return "financialRevenue"
    if "depreciacao" in account or "depreciacao" in category:
        return "depreciation"
    if "irrf" in account or "irrf" in category:
        return "irrf"
    if "irpj" in account or "irpj" in category:
        return "irpj"
    if "csll" in account or "csll" in category:
        return "csll"
    if "receita bruta" in account:
        return "gross"
    if "impostos" in account:
        return "taxes"
    if "deducoes" in account:
        return "deductions"
    if "custo dos servicos" in account or account.startswith("02. custos") or "servicos agregados" in category:
        return "costs"
    if "outras receitas" in account:
        return "otherRevenue"
    if "despesas administrativas" in account:
        return "admin"
    if "despesas com pessoal" in account:
        return "people"
    if "vendas e marketing" in account:
        return "sales"
    if "outros tributos" in account:
        return "otherTaxes"
    if "receitas financeiras" in account:
        return "financialRevenue"
    if "despesas financeiras" in account:
        return "financialExpense"
    return "other"


def default_percent_for_category(category):
    defaults = {
        norm("COFINS (Serviço)"): -0.0300,
        norm("CSLL (Serviço)"): -0.0108,
        norm("IRRF (Serviço)"): -0.0120,
        norm("ISS"): -0.0500,
        norm("ISS Retido"): -0.0500,
        norm("ISS (Retido)"): -0.0500,
        norm("PIS (Serviço)"): -0.0065,
    }
    return defaults.get(norm(category), 1.0)


def finalized_finance(bucket):
    gross = bucket.get("gross", 0)
    deductions = bucket.get("taxes", 0) + bucket.get("deductions", 0)
    net = gross + deductions
    costs = bucket.get("costs", 0) + bucket.get("otherRevenue", 0)
    margin = net + costs
    expenses = bucket.get("admin", 0) + bucket.get("people", 0) + bucket.get("sales", 0) + bucket.get("otherTaxes", 0)
    ebitda = margin + expenses
    depreciation = bucket.get("depreciation", 0)
    ebit = ebitda + depreciation
    financial = bucket.get("financialRevenue", 0) + bucket.get("financialExpense", 0)
    ebt = ebit + financial
    result_taxes = bucket.get("irpj", 0) + bucket.get("irrf", 0) + bucket.get("csll", 0)
    result = ebt + result_taxes
    out = dict(bucket)
    out.update(
        deductions=deductions,
        netRevenue=net,
        costsTotal=costs,
        contributionMargin=margin,
        expensesTotal=expenses,
        ebitda=ebitda,
        depreciation=depreciation,
        ebit=ebit,
        financialResult=financial,
        ebt=ebt,
        resultTaxes=result_taxes,
        netResult=result,
    )
    return out


def compact_finance_rows(rows):
    dimensions = (
        "source",
        "period",
        "tipo",
        "grupo",
        "account",
        "category",
        "client",
        "rateio",
        "project",
        "sourceHub",
        "hub",
        "expt",
        "city",
        "department",
        "vehicleType",
        "fleetType",
        "fleetOwner",
        "costType",
        "campaign",
    )
    grouped = {}
    for row in rows:
        key = tuple(row.get(name, "") for name in dimensions)
        if key not in grouped:
            grouped[key] = {name: row.get(name, "") for name in dimensions}
            grouped[key]["value"] = 0.0
        grouped[key]["value"] += row.get("value", 0) or 0
    return [row for row in grouped.values() if abs(row.get("value", 0)) > 0.0001]


def compact_operation_rows(rows):
    dimensions = (
        "period",
        "client",
        "project",
        "sourceHub",
        "hub",
        "expt",
        "city",
        "department",
        "vehicleType",
        "fleetType",
        "fleetOwner",
    )
    measures = (
        "routes",
        "freightValue",
        "km",
        "shipped",
        "delivered",
        "performanceWeighted",
        "performanceWeight",
        "evidenced",
        "fleetRoutes",
        "aggregatedRoutes",
    )
    grouped = {}
    for row in rows:
        key = tuple(row.get(name, "") for name in dimensions)
        if key not in grouped:
            grouped[key] = {name: row.get(name, "") for name in dimensions}
            for measure in measures:
                grouped[key][measure] = 0.0
        for measure in measures:
            grouped[key][measure] += row.get(measure, 0) or 0
    return [row for row in grouped.values() if any(abs(row.get(measure, 0)) > 0.0001 for measure in measures)]


def ledger_rows(rows):
    keep = (
        "source",
        "period",
        "account",
        "category",
        "client",
        "project",
        "sourceHub",
        "hub",
        "expt",
        "department",
        "vehicleType",
        "fleetType",
        "costType",
        "party",
        "invoice",
        "sourceRow",
        "originalValue",
        "value",
    )
    grouped = {}
    dimensions = tuple(name for name in keep if name not in {"value"})
    for row in rows:
        if abs(row.get("value", 0) or 0) <= 0.0001:
            continue
        key = tuple(row.get(name, "") for name in dimensions)
        if key not in grouped:
            grouped[key] = {name: row.get(name, "") for name in dimensions}
            grouped[key]["value"] = 0.0
        grouped[key]["value"] += row.get("value", 0) or 0
    return [row for row in grouped.values() if abs(row.get("value", 0) or 0) > 0.0001]


def load_budget_campaigns(wb):
    if "Campanhas" not in wb.sheetnames:
        return []
    ws = wb["Campanhas"]
    idx = header_index(ws)
    campaigns = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        year = get_by_header(row, idx, "Ano")
        month = get_by_header(row, idx, "Mês", "Mes")
        factor = safe_float(get_by_header(row, idx, "Fator Sazonal", "Fator", "Sazonal"))
        if not year or not month or abs(factor) <= 0.0001:
            continue
        try:
            period = f"{int(float(year)):04d}-{int(float(month)):02d}"
        except Exception:
            continue
        campaigns.append({
            "period": period,
            "client": clean(get_by_header(row, idx, "Cliente"), ""),
            "project": clean(get_by_header(row, idx, "Projeto"), ""),
            "hub": canonical_location(clean(get_by_header(row, idx, "Unidade", "Hub"), "")),
            "name": clean(get_by_header(row, idx, "Campanha"), ""),
            "factor": factor,
            "note": clean(get_by_header(row, idx, "Observação", "Observacao"), ""),
        })
    return campaigns


def budget_campaign_multiplier(item, campaigns):
    multiplier = 1.0
    names = []
    for campaign in campaigns:
        if campaign["period"] != item["period"]:
            continue
        if campaign["client"] and norm(campaign["client"]) != norm(item.get("client")):
            continue
        if campaign["project"] and norm(campaign["project"]) != norm(item.get("project")):
            continue
        if campaign["hub"] and norm(campaign["hub"]) != norm(item.get("hub")):
            continue
        multiplier *= 1 + campaign["factor"]
        if campaign["name"]:
            names.append(campaign["name"])
    return multiplier, names


def budget_campaign_base(item):
    account = norm(item.get("account"))
    category = norm(item.get("category"))
    cost_type = norm(item.get("costType"))
    if "RECEITA" in account or "IMPOST" in account or "DEDUC" in account:
        return True
    if "SERVICOS AGREGADOS" in category or "SERVIÇOS AGREGADOS" in category:
        return True
    return cost_type in {"VARIAVEL", "VARIÁVEL"}


def load_budget_rows(actual_periods=None):
    if not BUDGET_FILE.exists():
        return [], []
    wb = load_workbook(BUDGET_FILE, read_only=True, data_only=True)
    try:
        if "Budget" not in wb.sheetnames:
            return [], []
        campaigns = load_budget_campaigns(wb)
        ws = wb["Budget"]
        idx = header_index(ws)
        rows = []
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            period = clean(get_by_header(row, idx, "Período", "Periodo"), "")
            if not period:
                year = get_by_header(row, idx, "Ano")
                month = get_by_header(row, idx, "Mês", "Mes")
                if year and month:
                    try:
                        period = f"{int(float(year)):04d}-{int(float(month)):02d}"
                    except Exception:
                        period = ""
            value = safe_float(get_by_header(row, idx, "Valor Budget", "Budget", "Valor"))
            if not period or abs(value) <= 0.0001:
                continue
            hub = canonical_location(clean(get_by_header(row, idx, "Unidade", "Hub"), "N/D"))
            expt = canonical_location(clean(get_by_header(row, idx, "Expt"), hub))
            item = {
                "source": "Budget",
                "period": period,
                "scenario": clean(get_by_header(row, idx, "Cenário", "Cenario"), "Budget"),
                "account": clean(get_by_header(row, idx, "Conta DRE"), "N/D"),
                "category": clean(get_by_header(row, idx, "Categoria"), "N/D"),
                "client": clean(get_by_header(row, idx, "Cliente"), "N/D"),
                "project": clean(get_by_header(row, idx, "Projeto"), "N/D"),
                "sourceHub": hub,
                "hub": hub,
                "expt": expt,
                "department": expt,
                "vehicleType": clean(get_by_header(row, idx, "Tipo"), "N/D"),
                "fleetType": clean(get_by_header(row, idx, "Frota"), "N/D"),
                "costType": clean(get_by_header(row, idx, "Tipo Custo"), ""),
                "note": clean(get_by_header(row, idx, "Observação", "Observacao"), ""),
                "sourceRow": excel_row,
                "value": value,
            }
            if budget_campaign_base(item):
                multiplier, campaigns_applied = budget_campaign_multiplier(item, campaigns)
                if abs(multiplier - 1.0) > 0.0001:
                    item["value"] = item["value"] * multiplier
                    item["campaign"] = ", ".join(campaigns_applied)
            rows.append(item)
        if rows and actual_periods:
            budget_periods = sorted({item["period"] for item in rows})
            budget_years = {period[:4] for period in budget_periods}
            comparable_actual_periods = {period for period in actual_periods if period[:4] in budget_years}
            missing_actual_periods = sorted(comparable_actual_periods - set(budget_periods))
            if missing_actual_periods and budget_periods:
                template_period = budget_periods[0]
                template_rows = [item for item in rows if item["period"] == template_period]
                for period in missing_actual_periods:
                    year, month = period.split("-")
                    for item in template_rows:
                        clone = dict(item)
                        clone["period"] = period
                        clone["scenario"] = f"{item.get('scenario') or 'Budget'} - base comparativa"
                        clone["note"] = "Base comparativa criada pela média 4M para permitir previsto x realizado."
                        clone["sourceRow"] = None
                        rows.append(clone)
    finally:
        wb.close()
    return compact_finance_rows(rows), campaigns


def is_truthy_flag(value):
    text = norm(value)
    return bool(text) and text not in {"NAO", "NÃO", "N", "NO", "FALSE", "0"}


def build_data():
    if SOURCE_FILE.exists():
        shutil.copy2(SOURCE_FILE, WORKBOOK)
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)

    client_map = {}
    rateio_projects = set()
    for row in wb["Auxiliar_Cliente"].iter_rows(min_row=2, values_only=True):
        project = clean(row[0], "")
        client = clean(row[1], "")
        if project and client:
            client_map[project] = client

    dre_map = defaultdict(list)
    tax_mappings = []
    seen_dre_map = set()
    razao_aux_ws = wb["Auxiliar_Razao"]
    razao_aux_headers = [norm(cell.value) for cell in next(razao_aux_ws.iter_rows(min_row=1, max_row=1))]
    razao_aux_idx = {header: index for index, header in enumerate(razao_aux_headers) if header}
    launch_col = razao_aux_idx.get("LANCAMENTO", razao_aux_idx.get("LANÇAMENTO", razao_aux_idx.get("CATEGORIA", 0)))
    category_col = razao_aux_idx.get("CATEGORIA", launch_col)
    tipo_col = razao_aux_idx.get("TIPO", 1)
    grupo_col = razao_aux_idx.get("GRUPO", 2)
    account_col = razao_aux_idx.get("CONTA DO DRE", 3)
    percent_col = razao_aux_idx.get("ALIQUOTA", razao_aux_idx.get("ALÍQUOTA"))
    cost_type_col = razao_aux_idx.get("TIPO_CUSTO", razao_aux_idx.get("TIPO CUSTO"))
    fleet_col = razao_aux_idx.get("FROTA")
    fleet_categories = set()
    fleet_launches = set()
    cost_type_by_category = {}
    cost_type_by_launch = {}
    for row in razao_aux_ws.iter_rows(min_row=2, values_only=True):
        launch_key = clean(row[launch_col] if launch_col < len(row) else None, "")
        category = clean(row[category_col] if category_col < len(row) else None, launch_key)
        if not launch_key:
            continue
        percent_raw = row[percent_col] if percent_col is not None and percent_col < len(row) else None
        cost_type = clean(row[cost_type_col] if cost_type_col is not None and cost_type_col < len(row) else None, "")
        is_fleet_utility = is_truthy_flag(row[fleet_col]) if fleet_col is not None and fleet_col < len(row) else False
        item = {
            "tipo": clean(row[tipo_col] if tipo_col < len(row) else None),
            "grupo": clean(row[grupo_col] if grupo_col < len(row) else None),
            "account": clean(row[account_col] if account_col < len(row) else None),
            "category": category,
            "percent": safe_float(percent_raw) if percent_raw not in (None, "") else default_percent_for_category(category),
            "isFleetUtility": is_fleet_utility,
            "costType": cost_type,
        }
        if cost_type:
            cost_type_by_launch[norm(launch_key)] = cost_type
            cost_type_by_category[norm(category)] = cost_type
        if is_fleet_utility:
            fleet_launches.add(norm(launch_key))
            fleet_categories.add(norm(category))
        fingerprint = (norm(launch_key), norm(category), item["tipo"], item["grupo"], item["account"], item["percent"], item["isFleetUtility"], item["costType"])
        if fingerprint in seen_dre_map:
            continue
        seen_dre_map.add(fingerprint)
        dre_map[norm(launch_key)].append(item)
        if norm(item["account"]) == norm("02. Impostos") and norm(category) != norm("ISS Retido") and norm(category) != norm("ISS (Retido)"):
            tax_mappings.append(item)

    hub_map = {}
    fleet_plates = set()
    fleet_map = {}
    iss_rate_by_hub = {}
    hub_ws = wb["Auxiliar_Hub"]
    hub_headers = [norm(cell.value) for cell in next(hub_ws.iter_rows(min_row=1, max_row=1))]
    hub_idx = {header: index for index, header in enumerate(hub_headers) if header}
    plate_indexes = [index for index, header in enumerate(hub_headers) if "PLACA" in header]
    hub_code_names = {}
    unit_name_by_code = {}
    expt_name_by_code = {}
    hub_rows = list(hub_ws.iter_rows(min_row=2, values_only=True))
    for row in hub_rows:
        hub_rate_name = clean(row[hub_idx.get("TOTAL", 0)], "")
        hub_iss_rate = safe_float(row[hub_idx.get("ISS", 0)]) if "ISS" in hub_idx else 0
        if hub_rate_name and hub_iss_rate:
            iss_rate_by_hub[norm(hub_rate_name)] = hub_iss_rate

        code = clean(row[hub_idx.get("CODIGO", 0)], "")
        if code:
            code_hub_name = canonical_location(clean(row[hub_idx.get("NOME_HUB", hub_idx.get("CODIGO", 0))], code))
            code_expt_name = canonical_location(clean(row[hub_idx.get("NOME_EXPT", hub_idx.get("CODIGO", 0))], code))
            hub_code_names[norm(code)] = {"hubName": code_hub_name, "exptName": code_expt_name}
            unit_name_by_code[norm(code)] = code_hub_name
            expt_name_by_code[norm(code)] = code_expt_name

    for row in hub_rows:
        operation = clean(row[hub_idx.get("OPERACAO", 0)], "")
        origin_hub = canonical_location(clean(row[hub_idx.get("HUB", 1)], ""))
        city = clean(row[hub_idx.get("CIDADE", 2)], "")
        expt_code = clean(row[hub_idx.get("EXPT", 3)], "")
        if expt_code and origin_hub:
            unit_name_by_code.setdefault(norm(expt_code), origin_hub)
        code_names = hub_code_names.get(norm(expt_code), {})
        hub_name = canonical_location(code_names.get("hubName", origin_hub))
        expt_name = canonical_location(code_names.get("exptName", expt_code or hub_name))
        if code_like(expt_name) and origin_hub:
            expt_name = origin_hub
        if operation and origin_hub and city:
            hub_map[(norm(operation), norm(origin_hub), norm(city))] = {
                "hubName": hub_name,
                "expt": expt_name,
                "exptName": expt_name,
                "department": expt_name,
            }
        for index in plate_indexes:
            plate = clean(row[index], "")
            if plate:
                fleet_plates.add(norm(plate))
                fleet_value = clean(row[hub_idx.get("FROTA", index)], "Frota")
                fleet_map[norm(plate)] = fleet_value

    excluded_razao_categories = {
        norm("Clientes - Serviços Prestados"),
        norm("COFINS (Serviço)"),
        norm("CSLL (Serviço)"),
        norm("IRRF (Serviço)"),
        norm("ISS"),
        norm("ISS (Retido)"),
        norm("ISS Retido"),
        norm("PIS (Serviço)"),
        norm("Serviços Agregados"),
    }

    finance = defaultdict(lambda: defaultdict(float))
    finance_rows = []
    unified_rows = []
    allocation_basis = defaultdict(lambda: defaultdict(float))
    scoped_type_basis = defaultdict(lambda: defaultdict(float))

    operations = defaultdict(lambda: defaultdict(float))
    operation_rows = []
    ws = wb["Faturamento_Pagamento"]
    fp_idx = header_index(ws)
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        dt = get_by_header(row, fp_idx, "DATA")
        period = period_from_parts(
            get_by_header(row, fp_idx, "Ano Competência", "Ano Competencia"),
            get_by_header(row, fp_idx, "Mês Competência", "Mes Competencia"),
        )
        if not period and isinstance(dt, datetime):
            period = f"{dt.year}-{dt.month:02d}"
        if not period:
            continue
        launch = clean(get_by_header(row, fp_idx, "LANÇAMENTO", "LANCAMENTO"))
        plate = clean(get_by_header(row, fp_idx, "PLACA"), "")
        vehicle_type = clean(get_by_header(row, fp_idx, "TIPO"))
        project = clean(get_by_header(row, fp_idx, "OPERACAO", "OPERAÇÃO"))
        client = client_map.get(project, project)
        is_rateio = project in rateio_projects
        source_hub = canonical_location(clean(get_by_header(row, fp_idx, "HUB")))
        city = clean(get_by_header(row, fp_idx, "CIDADE"))
        hub_info = hub_map.get((norm(project), norm(source_hub), norm(city)), {})
        hub = canonical_location(hub_info.get("hubName", source_hub))
        expt = canonical_location(hub_info.get("expt", hub))
        department = canonical_location(hub_info.get("department", expt))
        fleet_type = "Frota" if plate and norm(plate) in fleet_plates else "Agregado"
        fleet_owner = fleet_map.get(norm(plate), "Agregado")
        value = safe_float(get_by_header(row, fp_idx, "VALOR"))
        invoice = clean(get_by_header(row, fp_idx, "INVOICE", "NF"), "")
        shipped = safe_float(get_by_header(row, fp_idx, "EMBARCADOS", "EMBARCADO"))
        delivered = safe_float(get_by_header(row, fp_idx, "ENTREGUES", "ENTREGUE"))
        performance = delivered / shipped if shipped else safe_float(get_by_header(row, fp_idx, "PERFORMANCE"))
        evidenced = 1 if norm(get_by_header(row, fp_idx, "EVIDENCIADO")) == "SIM" else 0

        is_route = norm(launch) in {"ROTA", "ROTAS"}
        op_values = {
            "routes": 1 if is_route else 0,
            "freightValue": value,
            "km": safe_float(get_by_header(row, fp_idx, "KM")),
            "shipped": shipped if is_route else 0,
            "delivered": delivered if is_route else 0,
            "performanceWeighted": performance * shipped if is_route else 0,
            "performanceWeight": shipped if is_route else 0,
            "evidenced": evidenced,
            "fleetRoutes": 1 if is_route and fleet_type == "Frota" else 0,
            "aggregatedRoutes": 1 if is_route and fleet_type == "Agregado" else 0,
        }
        op_keys = [
            ("all", period, "Todos", "Todos", "Todos"),
            ("client", period, client, "Todos", "Todos"),
            ("project", period, client, project, "Todos"),
            ("hub", period, client, project, hub),
        ]
        for key in op_keys:
            add_bucket(operations, key, op_values)
        operation_item = {
                "period": period,
                "date": dt.isoformat() if isinstance(dt, datetime) else "",
                "launch": launch,
                "client": client,
                "project": project,
                "sourceHub": source_hub,
                "hub": hub,
                "expt": expt,
                "city": city,
                "department": department,
                "plate": plate,
                "vehicleType": vehicle_type,
                "fleetType": fleet_type,
                "fleetOwner": fleet_owner,
                **op_values,
            }
        operation_rows.append(operation_item)
        if value and not is_rateio:
            alloc_key = allocation_key(operation_item)
            alloc_value = abs(value)
            allocation_basis[period][alloc_key] += alloc_value
            for scoped_key in scoped_allocation_keys(period, client, project, hub, expt):
                scoped_type_basis[scoped_key][alloc_key] += alloc_value

        mappings = dre_map.get(norm(launch), [])
        if not mappings:
            mappings = [{"tipo": "1. Lucro Bruto", "grupo": "02. Custos", "account": launch, "percent": -1.0}]
        expanded_mappings = []
        for mapping in mappings:
            expanded_mappings.append(mapping)
            if norm(mapping["account"]) == norm("01. Receita Bruta de Vendas"):
                expanded_mappings.extend(tax_mappings)

        for mapping in expanded_mappings:
            account = mapping["account"]
            category = mapping.get("category", launch)
            percent = mapping["percent"]
            is_fleet_utility = mapping.get("isFleetUtility", False)
            cost_type = mapping.get("costType", "")
            if norm(category) == norm("ISS"):
                percent = iss_rate_by_hub.get(norm(hub), percent)
            dre_value = value * percent
            field = classify_finance(account, category)
            display_tipo = mapping["tipo"]
            display_grupo = mapping["grupo"]
            display_account = account
            if field in {"irpj", "irrf", "csll"}:
                display_tipo = "4. Resultado Líquido"
                display_grupo = "Impostos antes do Resultado"
                display_account = category
            if field == "financialRevenue" and norm(category) == norm("Rendimentos de Aplicações"):
                display_tipo = "3. Resultado Financeiro"
                display_grupo = "02. Receitas Financeiras"
                display_account = "02. Receitas Financeiras"
            values = {field: dre_value}
            keys = [
                ("all", period, "Todos", "Todos", "Todos"),
                ("client", period, client, "Todos", "Todos"),
                ("project", period, client, project, "Todos"),
                ("hub", period, client, project, hub),
            ]
            for key in keys:
                add_bucket(finance, key, values)
            item = {
                "source": "Faturamento_Pagamento",
                "period": period,
                "date": dt.isoformat() if isinstance(dt, datetime) else "",
                "tipo": display_tipo,
                "grupo": display_grupo,
                "account": display_account,
                "category": category,
                "costType": cost_type,
                "value": dre_value,
                "originalValue": value,
                "party": client,
                "invoice": invoice,
                "sourceRow": excel_row,
                "client": client,
                "rateio": is_rateio,
                "project": project,
                "sourceHub": source_hub,
                "hub": hub,
                "city": city,
                "expt": expt,
                "department": department,
                "plate": plate,
                "vehicleType": "UTILITARIO" if is_fleet_utility else vehicle_type,
                "fleetType": "Frota" if is_fleet_utility else fleet_type,
                "fleetOwner": "Propria" if is_fleet_utility else fleet_owner,
            }
            finance_rows.append(item)
            unified_rows.append(item)

    ws = wb["Razão"]
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        period = period_from_parts(row[5], row[6])
        if not period:
            continue
        value = safe_float(row[7])
        if abs(value) < 0.0001:
            continue
        project = clean(row[11])
        client = clean(row[18] if len(row) > 18 else None, client_map.get(project, client_map.get(project, "Outros")))
        if client == "TOTAL":
            client = "REDEFRETE"
        raw_hub = clean(row[13])
        hub = canonical_location(unit_name_by_code.get(norm(raw_hub), raw_hub))
        expt = canonical_location(expt_name_by_code.get(norm(raw_hub), hub))
        department = canonical_location(clean(row[12]))
        account = clean(row[2])
        category = clean(row[3])
        party = clean(row[8])
        invoice = clean(row[10])
        cost_type = cost_type_by_category.get(norm(category), cost_type_by_launch.get(norm(category), ""))
        financeiro_fuel_expense = norm(category) == norm("Combustível Frota") and norm(hub) == norm("Financeiro")
        if financeiro_fuel_expense:
            account = "02. Despesas Administrativas"
            category = "Combustível"
            cost_type = ""
        if norm(category) in excluded_razao_categories:
            continue
        field = classify_finance(account, category)
        display_tipo = clean(row[0])
        display_grupo = clean(row[1])
        display_account = account
        if field in {"irpj", "irrf", "csll"}:
            display_tipo = "4. Resultado Líquido"
            display_grupo = "Impostos antes do Resultado"
            display_account = category
        if field == "financialRevenue" and norm(category) == norm("Rendimentos de Aplicações"):
            display_tipo = "3. Resultado Financeiro"
            display_grupo = "02. Receitas Financeiras"
            display_account = "02. Receitas Financeiras"
        is_fleet_utility = False if financeiro_fuel_expense else norm(category) in fleet_categories or norm(category) in fleet_launches
        is_razao_rateio = norm(department) in RATEIO_DEPARTMENTS
        global_rateio_base = allocation_basis.get(period, {})
        rateio_base = scoped_rateio_base(
            period,
            client,
            project,
            hub,
            expt,
            scoped_type_basis,
            global_rateio_base,
            is_fleet_utility,
        )
        if is_razao_rateio and rateio_base:
            base_total = sum(rateio_base.values())
            if base_total:
                for alloc_key, alloc_value in rateio_base.items():
                    (
                        alloc_client,
                        alloc_project,
                        alloc_hub,
                        alloc_expt,
                        alloc_source_hub,
                        alloc_city,
                        alloc_department,
                        alloc_vehicle_type,
                        alloc_fleet_type,
                        alloc_fleet_owner,
                    ) = alloc_key
                    allocated_value = value * (alloc_value / base_total)
                    values = {field: allocated_value}
                    keys = [
                        ("all", period, "Todos", "Todos", "Todos"),
                        ("client", period, alloc_client, "Todos", "Todos"),
                        ("project", period, alloc_client, alloc_project, "Todos"),
                        ("hub", period, alloc_client, alloc_project, alloc_hub),
                    ]
                    for key in keys:
                        add_bucket(finance, key, values)
                    item = {
                        "source": "Razao Rateado",
                        "period": period,
                        "date": row[9].isoformat() if isinstance(row[9], datetime) else "",
                        "tipo": display_tipo,
                        "grupo": display_grupo,
                        "account": display_account,
                        "category": category,
                        "costType": cost_type,
                        "value": allocated_value,
                        "originalValue": value,
                        "party": party,
                        "invoice": invoice,
                        "sourceRow": excel_row,
                        "client": alloc_client,
                        "rateio": True,
                        "rateioOriginProject": project,
                        "rateioOriginDepartment": department,
                        "project": alloc_project,
                        "sourceHub": alloc_source_hub,
                        "hub": alloc_hub,
                        "city": alloc_city,
                        "expt": alloc_expt,
                        "department": alloc_department,
                        "plate": "",
                        "vehicleType": "UTILITARIO" if is_fleet_utility else alloc_vehicle_type,
                        "fleetType": "Frota" if is_fleet_utility else alloc_fleet_type,
                        "fleetOwner": "Propria" if is_fleet_utility else alloc_fleet_owner,
                    }
                    finance_rows.append(item)
                    unified_rows.append(item)
                continue
        local_type_base = {}
        if not is_fleet_utility:
            for scoped_key in scoped_allocation_keys(period, client, project, hub, expt):
                local_type_base = scoped_type_basis.get(scoped_key, {})
                if local_type_base:
                    break
        if not is_fleet_utility and local_type_base:
            base_total = sum(local_type_base.values())
            if base_total:
                for alloc_key, alloc_value in local_type_base.items():
                    (
                        alloc_client,
                        alloc_project,
                        alloc_hub,
                        alloc_expt,
                        alloc_source_hub,
                        alloc_city,
                        alloc_department,
                        alloc_vehicle_type,
                        alloc_fleet_type,
                        alloc_fleet_owner,
                    ) = alloc_key
                    allocated_value = value * (alloc_value / base_total)
                    values = {field: allocated_value}
                    keys = [
                        ("all", period, "Todos", "Todos", "Todos"),
                        ("client", period, alloc_client, "Todos", "Todos"),
                        ("project", period, alloc_client, alloc_project, "Todos"),
                        ("hub", period, alloc_client, alloc_project, alloc_hub),
                    ]
                    for key in keys:
                        add_bucket(finance, key, values)
                    item = {
                        "source": "Razao Alocado Tipo",
                        "period": period,
                        "date": row[9].isoformat() if isinstance(row[9], datetime) else "",
                        "tipo": display_tipo,
                        "grupo": display_grupo,
                        "account": display_account,
                        "category": category,
                        "costType": cost_type,
                        "value": allocated_value,
                        "originalValue": value,
                        "party": party,
                        "invoice": invoice,
                        "sourceRow": excel_row,
                        "client": alloc_client,
                        "rateio": False,
                        "typeAllocationOriginProject": project,
                        "typeAllocationOriginDepartment": department,
                        "project": alloc_project,
                        "sourceHub": alloc_source_hub,
                        "hub": alloc_hub,
                        "city": alloc_city,
                        "expt": alloc_expt,
                        "department": alloc_department,
                        "plate": "",
                        "vehicleType": alloc_vehicle_type,
                        "fleetType": alloc_fleet_type,
                        "fleetOwner": alloc_fleet_owner,
                    }
                    finance_rows.append(item)
                    unified_rows.append(item)
                continue
        values = {field: value}
        keys = [
            ("all", period, "Todos", "Todos", "Todos"),
            ("client", period, client, "Todos", "Todos"),
            ("project", period, client, project, "Todos"),
            ("hub", period, client, project, hub),
        ]
        for key in keys:
            add_bucket(finance, key, values)
        item = {
            "source": "Razao",
            "period": period,
            "date": row[9].isoformat() if isinstance(row[9], datetime) else "",
            "tipo": display_tipo,
            "grupo": display_grupo,
            "account": display_account,
            "category": category,
            "costType": cost_type,
            "value": value,
            "originalValue": value,
            "party": party,
            "invoice": invoice,
            "sourceRow": excel_row,
            "client": client,
            "rateio": is_razao_rateio,
            "project": project,
            "sourceHub": raw_hub,
            "hub": hub,
            "city": "N/D",
            "expt": expt,
            "department": department,
            "plate": "",
            "vehicleType": "UTILITARIO" if is_fleet_utility else "N/D",
            "fleetType": "Frota" if is_fleet_utility else "N/D",
            "fleetOwner": "Propria" if is_fleet_utility else "N/D",
        }
        finance_rows.append(item)
        unified_rows.append(item)

    compact_unified_rows = compact_finance_rows(unified_rows)
    compact_operation_rows_out = compact_operation_rows(operation_rows)

    finance_out = []
    for (level, period, client, project, hub), values in finance.items():
        item = finalized_finance(values)
        item.update(level=level, period=period, client=client, project=project, hub=hub)
        finance_out.append(item)

    operations_out = []
    for (level, period, client, project, hub), values in operations.items():
        weight = values.get("performanceWeight", 0)
        performance = values.get("performanceWeighted", 0) / weight if weight else 0
        item = dict(values)
        item.update(level=level, period=period, client=client, project=project, hub=hub, performance=performance)
        operations_out.append(item)

    periods = sorted({item["period"] for item in finance_out} | {item["period"] for item in operations_out})
    clients = sorted({item["client"] for item in finance_out + operations_out if item["client"] != "Todos"})
    projects = sorted({item["project"] for item in finance_out + operations_out if item["project"] != "Todos"})
    hubs = sorted({item["hub"] for item in finance_out + operations_out if item["hub"] != "Todos"})
    units = unique_labels(unit_value(item) for item in compact_unified_rows + compact_operation_rows_out if unit_value(item) and unit_value(item) != "N/D")
    expts = unique_labels(item.get("expt") for item in compact_unified_rows + compact_operation_rows_out if item.get("expt") and item.get("expt") != "N/D")
    vehicle_types = sorted({item["vehicleType"] for item in compact_unified_rows + compact_operation_rows_out if item.get("vehicleType") and item["vehicleType"] != "N/D"})
    fleet_types = sorted({item["fleetType"] for item in compact_unified_rows + compact_operation_rows_out if item.get("fleetType") and item["fleetType"] != "N/D"})
    budget_rows, budget_campaigns = load_budget_rows(periods)
    budget_periods = sorted({item["period"] for item in budget_rows})
    budget_scenarios = sorted({item.get("scenario") for item in budget_rows if item.get("scenario")})

    data = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "source": str(SOURCE_FILE),
            "periods": periods,
            "clients": clients,
            "projects": projects,
            "hubs": hubs,
            "units": units,
            "expts": expts,
            "departments": units,
            "vehicleTypes": vehicle_types,
            "fleetTypes": fleet_types,
            "budgetPeriods": budget_periods,
            "budgetScenarios": budget_scenarios,
            "budgetRows": len(budget_rows),
            "budgetCampaigns": len(budget_campaigns),
            "financeRows": len(finance_rows),
            "unifiedRowsRaw": len(unified_rows),
            "unifiedRows": len(compact_unified_rows),
            "operationRows": len(operation_rows),
            "operationRowsCompact": len(compact_operation_rows_out),
            "fleetPlateSourceCount": len(fleet_plates),
            "sourceMtimeNs": SOURCE_FILE.stat().st_mtime_ns if SOURCE_FILE.exists() else WORKBOOK.stat().st_mtime_ns,
        },
        "finance": finance_out,
        "operations": operations_out,
        "operationRows": compact_operation_rows_out,
        "unifiedRows": compact_unified_rows,
        "budgetRows": budget_rows,
    }
    ledger = ledger_rows(unified_rows)
    pl_database.write_database(data, ledger=ledger)
    if os.environ.get("PL_SKIP_LEGACY_JSON") != "1":
        DATA_OUT.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        LEDGER_OUT.write_text(json.dumps(ledger, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return data


CSS = r"""
:root{--blue:#0d5be1;--green:#0aa36f;--red:#ef4444;--violet:#7c3aed;--teal:#0891b2;--ink:#061a46;--muted:#63708d;--line:#dce5f3;--panel:#fff;--bg:#eef2f8;--shadow:0 10px 26px rgba(9,24,66,.07)}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);font-family:Inter,Segoe UI,Arial,sans-serif;color:var(--ink);font-weight:500}
.app{display:grid;grid-template-columns:230px 1fr;min-height:100vh}
.sidebar{background:#071f52;color:#fff;padding:24px 14px;display:flex;flex-direction:column;gap:24px}
.brand-logo{width:168px}.nav{display:grid;gap:8px}
.nav button{height:40px;border:0;border-radius:7px;background:transparent;color:#fff;display:flex;align-items:center;gap:11px;padding:0 13px;font-size:13px;font-weight:700;cursor:pointer}
.nav button.active,.nav button:hover{background:var(--blue)}
.nav svg,.refresh svg{width:17px;height:17px;fill:none;stroke:currentColor;stroke-width:2}
.main{padding:14px 20px;overflow:hidden}
.topbar{display:grid;gap:8px;margin-bottom:10px}
h1{margin:0;font-size:23px;line-height:1.04}
.filter-card{display:grid;grid-template-columns:repeat(10,minmax(0,1fr));gap:6px;align-items:center;background:#fff;border:1px solid var(--line);border-radius:8px;box-shadow:var(--shadow);padding:7px}
.filter-card .control,.filter-card .month-filter{width:100%;min-width:0}
.control{height:31px;min-width:0;border:1px solid var(--line);border-radius:7px;background:#fff;display:flex;align-items:center;gap:5px;padding:0 8px}
.control label{color:var(--muted);font-size:9.8px;font-weight:650}
.control select{border:0;outline:0;background:#fff;font-weight:650;color:var(--ink);min-width:0;max-width:100%;font-size:11.5px}
.month-filter{position:relative;height:31px;min-width:0}
.month-filter summary{list-style:none;cursor:pointer;display:flex;align-items:center;gap:5px;height:100%;font-size:11.5px;font-weight:650;white-space:nowrap}
.month-filter summary::-webkit-details-marker{display:none}
.month-filter .check-list{position:absolute;z-index:20;top:40px;left:0;min-width:210px;max-height:250px;overflow:auto;background:#fff;border:1px solid var(--line);border-radius:8px;box-shadow:var(--shadow);padding:7px}
.check-list label{display:flex;gap:8px;align-items:center;padding:5px 7px;color:var(--ink);font-size:12px;white-space:nowrap}
.refresh{height:31px;border:0;border-radius:7px;background:var(--blue);color:#fff;padding:0 10px;display:flex;align-items:center;justify-content:center;gap:6px;font-size:12px;font-weight:750;cursor:pointer;white-space:nowrap}
#clearFilters{height:31px;border:1px solid var(--line);border-radius:7px;background:#fff;color:var(--muted);padding:0 8px;font-size:10px;font-weight:650;cursor:pointer;white-space:nowrap}
.budget-actions{display:flex;justify-content:flex-end;align-items:center;gap:8px;margin:-2px 0 8px}
.budget-refresh{height:27px;border:1px solid rgba(13,91,225,.24);border-radius:7px;background:#fff;color:var(--blue);padding:0 9px;font-size:10.5px;font-weight:700;cursor:pointer;white-space:nowrap}
.budget-refresh:hover{background:#f7fbff;border-color:rgba(13,91,225,.42)}
.budget-status{color:var(--muted);font-size:10.5px;font-weight:650}
.view{display:none}.view.active{display:block}
.kpis{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:9px;margin-bottom:9px}
.card,.panel{position:relative;background:#fff;border:1px solid var(--line);border-radius:7px;box-shadow:var(--shadow);overflow:hidden}
.card:before,.panel:before{content:"";position:absolute;inset:0 0 auto;height:4px;background:var(--accent,var(--blue))}
.kpi{padding:10px 15px;min-height:66px}
.kpi-label{font-size:10px;text-transform:uppercase;color:var(--muted);font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.kpi-value{font-size:20px;line-height:1.02;margin-top:4px;font-weight:800}
.kpi-note{font-size:10.5px;color:var(--muted);margin-top:3px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.grid{display:grid;grid-template-columns:repeat(12,minmax(0,1fr));gap:10px}
.grid>.panel{grid-column:span 4}.grid>.panel.wide{grid-column:span 7}.grid>.panel:nth-child(2){grid-column:span 5}
.panel{padding:10px 12px}.panel.full{grid-column:1/-1}
h2{font-size:14px;line-height:1.1;margin:0 0 8px;font-weight:700}
.chart-wrap{height:224px}.panel.wide .chart-wrap{height:360px}
.table-wrap{max-height:360px;overflow:auto;border:1px solid var(--line);border-radius:7px}
.table{width:100%;border-collapse:collapse;font-size:11.5px}
.table th,.table td{padding:6px 10px;border-bottom:1px solid var(--line);text-align:right;vertical-align:middle}
.table th:first-child,.table td:first-child{text-align:left}
.table th{background:#f8fbff;color:var(--muted);font-size:10px;text-transform:uppercase;position:sticky;top:0;z-index:1;font-weight:650}
.table .sub{display:block;margin-top:2px;color:var(--muted);font-size:9px;font-weight:550}
.ops-section td{background:#f8fbff;color:var(--muted);font-size:10px;text-transform:uppercase;font-weight:750;letter-spacing:.02em}
.ops-total td{background:#fbfdff;font-weight:750}
.pl-table-wrap .table{font-size:11.5px}
.pl-table-wrap .table th:first-child,.pl-table-wrap .table td:first-child{min-width:280px;font-size:12px}
.pl-table-wrap .table td:not(:first-child){font-size:12px;font-weight:650}
.pl-table-wrap .table .sub{font-size:9.5px}
.unit-pl-table{table-layout:fixed;min-width:max-content}
.unit-pl-table th,.unit-pl-table td{text-align:center!important;min-width:132px;width:132px}
.unit-pl-table th:first-child,.unit-pl-table td:first-child{text-align:left!important;min-width:280px;width:280px}
.total-row td{font-weight:750;background:#fbfdff}.group-row td{font-weight:600}.detail-row td{background:#fff;font-weight:500;color:#294164}
.detail-label{padding-left:26px}.drill-btn{width:17px;height:17px;border:0;background:transparent;color:var(--blue);font-weight:950;cursor:pointer;margin-right:4px}
.value-pos{color:#009b63}.value-neg{color:#e52e34}
.chart{width:100%;height:100%;display:block}.legend{display:flex;gap:14px;flex-wrap:wrap;margin-top:-22px;color:var(--muted);font-size:11px}.legend span{display:inline-flex;align-items:center;gap:6px}
.sw{width:22px;height:8px;border-radius:99px}
.rank{display:grid;gap:7px}.rank-row{display:grid;grid-template-columns:minmax(90px,1fr) 88px;gap:9px;align-items:center;font-size:12px}
.track{grid-column:1/-1;height:7px;background:#edf2fa;border-radius:99px;overflow:hidden}.fill{height:100%;background:var(--blue);border-radius:inherit}
.ops-rank-grid{position:relative;display:block;min-height:640px}
.ops-rank-grid .panel{position:absolute;display:flex;flex-direction:column;min-height:220px;min-width:360px;resize:both;overflow:auto}
.ops-rank-grid .panel h2{cursor:grab;user-select:none;touch-action:none}
.ops-rank-grid .panel h2:active{cursor:grabbing}
.ops-rank-grid .panel.dragging{opacity:.72;outline:2px dashed rgba(13,91,225,.38);box-shadow:0 18px 42px rgba(11,33,78,.18)}
.ops-rank-grid .panel>div[id$="Rank"]{display:flex;flex-direction:column;min-height:0;flex:1}
.ops-rank-grid .table-wrap{max-height:none;min-height:190px;overflow:auto;flex:1}
.sort-th{cursor:pointer;user-select:none;white-space:nowrap}
.sort-th .sort-arrow{margin-left:4px;color:var(--blue);font-size:11px}
.pl-module-grid{display:grid;gap:10px}.pl-table-wrap{max-height:calc(100vh - 278px)}
.panel-head{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:10px}.panel-head h2{margin:0}.panel-note{color:var(--muted);font-size:11px;font-weight:850;text-align:right}
.comparison-wrap{max-height:260px}
.note-btn{position:absolute;top:9px;right:9px;width:28px;height:28px;border:1px solid var(--line);border-radius:7px;background:rgba(255,255,255,.94);color:var(--muted);display:inline-grid;place-items:center;cursor:pointer;box-shadow:0 8px 18px rgba(20,44,88,.08);z-index:3}
.note-btn svg{width:15px;height:15px;fill:none;stroke:currentColor;stroke-width:2}
.note-btn:hover,.note-btn.active{border-color:rgba(13,91,225,.35);background:#f7fbff;color:var(--blue)}
.note-btn.active:after{content:"";position:absolute;top:5px;right:5px;width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 0 2px #fff}
.kpi{padding-right:48px}.panel h2{padding-right:34px}
.note-modal-backdrop{position:fixed;inset:0;background:rgba(6,26,70,.32);display:flex;align-items:center;justify-content:center;padding:22px;z-index:50}
.note-modal-backdrop[hidden]{display:none}
.note-modal{width:min(620px,100%);background:#fff;border:1px solid var(--line);border-radius:10px;box-shadow:0 24px 70px rgba(9,24,66,.24);padding:16px}
.unit-modal{width:min(1180px,96vw);max-height:88vh;overflow:hidden;display:flex;flex-direction:column}
.ledger-modal{width:min(1280px,96vw);max-height:88vh;overflow:hidden;display:flex;flex-direction:column}
.unit-modal .note-modal-head{align-items:center}.unit-modal .note-btn{position:static;flex:0 0 auto}.unit-modal .note-close{flex:0 0 auto}
.ledger-modal .note-modal-head{align-items:center}.ledger-modal .note-close{flex:0 0 auto}
.note-modal-head{display:flex;align-items:flex-start;justify-content:space-between;gap:14px;margin-bottom:12px}
.note-modal h3{margin:0;font-size:18px;line-height:1.2}.note-modal p{margin:4px 0 0;color:var(--muted);font-size:12px}
.unit-modal-body{overflow:auto;border:1px solid var(--line);border-radius:8px}
.ledger-modal-body{overflow:auto;border:1px solid var(--line);border-radius:8px}
.ledger-table{min-width:760px}.ledger-table td:first-child,.ledger-table th:first-child{min-width:84px}.ledger-table td:nth-child(4){font-weight:650}
.ledger-dbl{cursor:zoom-in}.ledger-dbl:hover td{background:#f3f7ff}
.breakeven-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:8px;margin:4px 0 10px}
.be-card{border:1px solid var(--line);border-radius:8px;background:#fbfdff;padding:8px 10px;min-height:58px}
.be-label{font-size:9.5px;text-transform:uppercase;color:var(--muted);font-weight:750;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.be-value{font-size:17px;line-height:1.1;margin-top:4px;font-weight:800;color:var(--ink)}
.be-note{font-size:10px;color:var(--muted);margin-top:3px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.unit-modal .unit-modal-body{max-height:48vh}
.click-row{cursor:pointer}.click-row:hover td{background:#f3f7ff}
.unit-modal-actions{display:flex;justify-content:flex-end;gap:8px;margin-top:12px}
.unit-modal-actions button{height:32px;border-radius:7px;padding:0 12px;font-weight:700;cursor:pointer}
.unit-modal-actions .secondary{border:1px solid var(--line);background:#fff;color:var(--muted)}
.note-close{width:30px;height:30px;border:1px solid var(--line);border-radius:7px;background:#fff;color:var(--muted);font-weight:950;cursor:pointer}
#noteText{width:100%;height:210px;resize:vertical;border:1px solid var(--line);border-radius:8px;padding:12px;font:600 13px/1.45 Inter,Segoe UI,Arial,sans-serif;color:var(--ink);outline:none}
#noteText:focus{border-color:#82adff;box-shadow:0 0 0 3px rgba(13,91,225,.12)}
.note-actions{display:flex;justify-content:flex-end;gap:8px;margin-top:12px}.note-actions button{height:34px;border-radius:7px;padding:0 14px;font-weight:900;cursor:pointer}
.note-actions .secondary{border:1px solid var(--line);background:#fff;color:var(--muted)}.note-actions .primary{border:0;background:var(--blue);color:#fff}
@media(min-width:1450px){.main{padding:18px 22px}.panel.wide .chart-wrap{height:360px}.kpi{min-height:80px}}
@media(max-width:1450px){.filter-card{grid-template-columns:repeat(5,minmax(0,1fr))}}
@media(max-width:1200px){.app{grid-template-columns:86px 1fr}.brand-logo{width:54px}.nav span{display:none}.kpis,.grid{grid-template-columns:1fr 1fr}.grid>.panel,.grid>.panel:nth-child(2){grid-column:auto}.grid>.panel.wide{grid-column:1/-1}}
@media(max-width:900px){.ops-rank-grid .panel{min-width:280px}}
"""


JS = r"""
const DATA_URL='/data?scope=base&ts='+Date.now(),BUDGET_DATA_URL='/data?scope=budget&ts=';let DATA=null,NOTES={},activeNoteKey='',budgetDataLoaded=false;const POPUP_GROUPS=new Map(),LEDGER_CONTEXTS=new Map();const OPS_LAYOUT_KEY='plOpsCardLayout';const state={year:'',months:['all'],client:'all',project:'all',unit:'all',expt:'all',type:'all',fleet:'all',view:'dashboard',drill:{},rankSort:{unitRank:'desc',typeRank:'desc',fleetRank:'desc'}};const $=id=>document.getElementById(id);
const BRL=new Intl.NumberFormat('pt-BR',{style:'currency',currency:'BRL',maximumFractionDigits:0});const PCT=new Intl.NumberFormat('pt-BR',{style:'percent',minimumFractionDigits:1,maximumFractionDigits:1});
function fmtMoney(v){return BRL.format((Number(v)||0)/1000)}function fmtPct(v){return Number.isFinite(v)?PCT.format(v):'-'}function cls(v){return !Number.isFinite(v)||Math.abs(v)<.0001?'':v>=0?'value-pos':'value-neg'}
function esc(v){return String(v??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
function monthLabel(p){const [y,m]=p.split('-');return `${m}/${y}`}function shortMonth(p){const [y,m]=p.split('-');return `${m}/${y.slice(2)}`}
function years(){return [...new Set([...(DATA.meta.periods||[]),...(DATA.meta.budgetPeriods||[])].map(p=>p.slice(0,4)))].sort((a,b)=>b.localeCompare(a))}
function periodsForYear(y){const source=state.view==='budget'?[...(DATA.meta.periods||[]),...(DATA.meta.budgetPeriods||[])]:DATA.meta.periods;return [...new Set(source.filter(p=>p.startsWith(y+'-')))].sort()}
function displayLabel(v){const keep=new Set(['P&L','T.I','EDSP','FM','LM','LH','SD']);return String(v||'').split(/(\s+|-)/).map(part=>{if(!part.trim()||part==='-')return part;const raw=part.replace(/[.,]/g,'');if(keep.has(raw.toUpperCase()))return part.toUpperCase();return part.charAt(0).toUpperCase()+part.slice(1).toLocaleLowerCase('pt-BR')}).join('')}
function fillSelect(id,values,all,labelFn=displayLabel){const el=$(id);el.innerHTML=(all?`<option value="all">${all}</option>`:'')+values.map(v=>`<option value="${v}">${labelFn(v)}</option>`).join('')}
function allOptionRows(){return DATA.unifiedRows.concat(DATA.operationRows||[],state.view==='budget'?(DATA.budgetRows||[]):[])}
function codeLike(v){const s=String(v||'').trim();return !!s&&/\d/.test(s)&&s.length<=6&&/^[A-Za-z0-9]+$/.test(s)}
function unitValue(r){if(r.hub&&r.hub!=='N/D'&&!codeLike(r.hub))return r.hub;if(r.department&&r.department!=='N/D'&&!codeLike(r.department))return r.department;if(r.sourceHub&&r.sourceHub!=='N/D'&&!codeLike(r.sourceHub))return r.sourceHub;const fallback=r.hub||r.sourceHub||r.department;return codeLike(fallback)?'Unidade não mapeada':fallback}
function isOperationalRow(r){const unit=canonicalKey(unitValue(r)),hub=canonicalKey(r.hub),source=canonicalKey(r.sourceHub),client=canonicalKey(r.client),project=canonicalKey(r.project);if(client==='REDEFRETE'&&project==='REDEFRETE'&&(unit==='FINANCEIRO'||hub==='FINANCEIRO'||source==='SHARE'))return false;return unit&&!['FINANCEIRO','ADMINISTRACAO','ADMINISTRACAO','RECURSOSHUMANOS','CONTROLADORIA','JURIDICO','CONTABILIDADE','MARKETING','TI'].includes(unit)}
function optionValue(r,key){if(key==='client')return r.client;if(key==='project')return r.project;if(key==='unit')return unitValue(r);if(key==='expt')return r.expt;if(key==='type')return r.vehicleType;if(key==='fleet')return r.fleetType;return''}
function rowsForOption(key){const periods=currentPeriods();return allOptionRows().filter(r=>periods.includes(r.period)&&(key==='client'||state.client==='all'||r.client===state.client)&&(key==='client'||key==='project'||state.project==='all'||r.project===state.project)&&(key==='client'||key==='project'||key==='unit'||state.unit==='all'||canonicalKey(unitValue(r))===canonicalKey(state.unit))&&(key==='client'||key==='project'||key==='unit'||key==='expt'||state.expt==='all'||canonicalKey(r.expt)===canonicalKey(state.expt))&&(key==='client'||key==='project'||key==='unit'||key==='expt'||key==='type'||state.type==='all'||r.vehicleType===state.type))}
function canonicalKey(v){return String(v||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toUpperCase().replace(/[^A-Z0-9]/g,'')}
function preferredOption(values){const clean=values.filter(Boolean);return clean.find(v=>String(v)!==String(v).toUpperCase())||clean[0]||''}
function valuesForOption(key){const groups=new Map();for(const value of rowsForOption(key).map(r=>optionValue(r,key)).filter(v=>v&&v!=='Todos'&&v!=='N/D')){const k=canonicalKey(value);if(!groups.has(k))groups.set(k,[]);groups.get(k).push(value)}return [...groups.values()].map(preferredOption).sort((a,b)=>String(a).localeCompare(String(b),'pt-BR'))}
function syncSelect(id,key,values){if(!values.includes(state[key]))state[key]='all';fillSelect(id,values,'Todos');$(id).value=state[key]}
function noteIconSvg(){return '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M6 4h12v16H6z"></path><path d="M9 8h6M9 12h6M9 16h4"></path></svg>'}
async function loadNotes(){try{const r=await fetch(new URL('/notes',location.origin).href,{cache:'no-store'});NOTES=await r.json();if(!NOTES||typeof NOTES!=='object')NOTES={}}catch(err){try{NOTES=JSON.parse(localStorage.getItem('plNotes')||'{}')}catch(_){NOTES={}}}}
async function saveNotes(){localStorage.setItem('plNotes',JSON.stringify(NOTES));const r=await fetch(new URL('/notes',location.origin).href,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(NOTES),cache:'no-store'});if(!r.ok)throw new Error('Não foi possível salvar a nota no arquivo local.')}
function panelNoteKey(panel){const view=panel.closest('.view')?.id||'global';const title=panel.querySelector('h2')?.textContent?.trim()||'Card';return `${view}:${title}`}
function ensureNoteButtons(){document.querySelectorAll('.panel').forEach(panel=>{if(panel.querySelector(':scope > .note-btn'))return;const title=panel.querySelector('h2')?.textContent?.trim();if(!title)return;const btn=document.createElement('button');btn.type='button';btn.className='note-btn';btn.innerHTML=noteIconSvg();btn.title='Notas explicativas';btn.setAttribute('aria-label','Editar nota explicativa');btn.dataset.note=panelNoteKey(panel);btn.dataset.noteTitle=title;panel.appendChild(btn)});updateNoteButtons()}
function updateNoteButtons(){document.querySelectorAll('[data-note]').forEach(btn=>btn.classList.toggle('active',!!String(NOTES[btn.dataset.note]||'').trim()))}
function openNote(key,title){activeNoteKey=key;$('noteTitle').textContent=title||'Notas explicativas';$('noteText').value=NOTES[key]||'';$('noteModal').hidden=false;$('noteText').focus()}
function closeNote(){activeNoteKey='';$('noteModal').hidden=true}
async function saveCurrentNote(){const text=$('noteText').value.trim();if(text)NOTES[activeNoteKey]=text;else delete NOTES[activeNoteKey];try{await saveNotes();updateNoteButtons();closeNote()}catch(err){alert(err.message)}}
function openNoteByKey(key,title){openNote(key,title)}
function opsCardId(panel){return panel.querySelector('#unitRank')?'unitRank':panel.querySelector('#typeRank')?'typeRank':panel.querySelector('#fleetRank')?'fleetRank':''}
function readOpsLayout(){try{return JSON.parse(localStorage.getItem(OPS_LAYOUT_KEY)||'{}')}catch(_){return {}}}
function writeOpsLayout(layout){localStorage.setItem(OPS_LAYOUT_KEY,JSON.stringify(layout||{}))}
function defaultOpsBox(grid,i){const gap=14,w=Math.max(360,Math.floor((grid.clientWidth-gap)/2)),h=330;return {x:(i%2)*(w+gap),y:Math.floor(i/2)*(h+gap),w,h}}
function legacyOpsBox(grid,pos,i){if(!pos||!('col'in pos)||!('row'in pos))return defaultOpsBox(grid,i);const base=defaultOpsBox(grid,(Number(pos.row||1)-1)*2+(Number(pos.col||1)-1));return base}
function setOpsBox(panel,box){panel.style.left=Math.max(0,Number(box.x)||0)+'px';panel.style.top=Math.max(0,Number(box.y)||0)+'px';panel.style.width=Math.max(320,Number(box.w)||360)+'px';panel.style.height=Math.max(220,Number(box.h)||300)+'px'}
function getOpsBox(panel){return {x:parseFloat(panel.style.left)||0,y:parseFloat(panel.style.top)||0,w:panel.offsetWidth||360,h:panel.offsetHeight||300}}
function updateOpsBoardHeight(){const grid=document.querySelector('#view-ops .ops-rank-grid');if(!grid)return;let bottom=620;[...grid.querySelectorAll('.panel')].forEach(panel=>{const b=getOpsBox(panel);bottom=Math.max(bottom,b.y+b.h+18)});grid.style.minHeight=bottom+'px'}
function saveOpsLayout(){const grid=document.querySelector('#view-ops .ops-rank-grid');if(!grid)return;const layout={boxes:{}};[...grid.querySelectorAll('.panel')].forEach(panel=>{const id=opsCardId(panel);if(!id)return;layout.boxes[id]=getOpsBox(panel)});writeOpsLayout(layout);updateOpsBoardHeight()}
function setupOpsCards(){const grid=document.querySelector('#view-ops .ops-rank-grid');if(!grid)return;const layout=readOpsLayout();const panels=[...grid.querySelectorAll('.panel')];let resizeTimer=null;panels.forEach((panel,i)=>{const id=opsCardId(panel),handle=panel.querySelector('h2');panel.dataset.cardId=id;panel.draggable=false;if(layout.boxes&&layout.boxes[id])setOpsBox(panel,layout.boxes[id]);else setOpsBox(panel,legacyOpsBox(grid,layout.positions&&layout.positions[id],i));if(handle&&!handle.dataset.dragReady){handle.dataset.dragReady='1';handle.title='Arraste o título para mover este card livremente';let startX=0,startY=0,startBox=null,dragging=false;handle.onpointerdown=e=>{if(e.button!==0)return;startX=e.clientX;startY=e.clientY;startBox=getOpsBox(panel);dragging=false;handle.setPointerCapture(e.pointerId);panel.classList.add('dragging');panel.style.zIndex=20;e.preventDefault()};handle.onpointermove=e=>{if(!panel.classList.contains('dragging'))return;if(!dragging&&Math.hypot(e.clientX-startX,e.clientY-startY)<5)return;dragging=true;const rect=grid.getBoundingClientRect(),x=Math.max(0,Math.min(grid.clientWidth-panel.offsetWidth,startBox.x+e.clientX-startX)),y=Math.max(0,startBox.y+e.clientY-startY);panel.style.left=x+'px';panel.style.top=y+'px';updateOpsBoardHeight()};handle.onpointerup=e=>{if(handle.hasPointerCapture(e.pointerId))handle.releasePointerCapture(e.pointerId);panel.classList.remove('dragging');panel.style.zIndex='';saveOpsLayout()};handle.onpointercancel=e=>{if(handle.hasPointerCapture(e.pointerId))handle.releasePointerCapture(e.pointerId);panel.classList.remove('dragging');panel.style.zIndex='';saveOpsLayout()}}panel.onmouseup=saveOpsLayout;if(!panel.dataset.resizeWatch){panel.dataset.resizeWatch='1';new ResizeObserver(()=>{clearTimeout(resizeTimer);resizeTimer=setTimeout(saveOpsLayout,180)}).observe(panel)}});updateOpsBoardHeight()}
function refreshHierarchicalFilters(){syncSelect('clientFilter','client',valuesForOption('client'));syncSelect('projectFilter','project',valuesForOption('project'));syncSelect('unitFilter','unit',valuesForOption('unit'));syncSelect('exptFilter','expt',valuesForOption('expt'));syncSelect('typeFilter','type',valuesForOption('type'));syncSelect('fleetFilter','fleet',valuesForOption('fleet'))}
function clearFilters(){const ys=years();state.year=ys[0];state.months=['all'];state.client=state.project=state.unit=state.expt=state.type=state.fleet='all';$('yearFilter').value=state.year;fillMonthChecks();render()}
function fillMonthChecks(){const values=periodsForYear(state.year);$('monthChecks').innerHTML=`<label><input type="checkbox" value="all"> Todos</label>`+values.map(v=>`<label><input type="checkbox" value="${v}"> ${monthLabel(v)}</label>`).join('');applyMonthSelection();document.querySelectorAll('#monthChecks input').forEach(i=>i.onchange=e=>{const target=e.target;if(target.value==='all'){state.months=target.checked?['all']:[]}else{const selected=[...document.querySelectorAll('#monthChecks input:checked')].map(x=>x.value).filter(v=>v!=='all');state.months=selected.length===values.length||!selected.length?['all']:selected}applyMonthSelection();render()})}
function applyMonthSelection(){const all=state.months.includes('all');document.querySelectorAll('#monthChecks input').forEach(i=>i.checked=all||state.months.includes(i.value));$('monthSummary').textContent=all?'Todos':state.months.length===1?shortMonth(state.months[0]):'Diversos'}
function currentPeriods(){return state.months.includes('all')?periodsForYear(state.year):state.months}
function actualPeriodsForSelection(){const actual=new Set(DATA.meta.periods||[]);return currentPeriods().filter(p=>actual.has(p))}
function budgetComparePeriods(){const actual=actualPeriodsForSelection();return actual.length?actual:currentPeriods()}
function passFilters(r,periods){return periods.includes(r.period)&&(state.client==='all'||r.client===state.client)&&(state.project==='all'||r.project===state.project)&&(state.unit==='all'||canonicalKey(unitValue(r))===canonicalKey(state.unit))&&(state.expt==='all'||canonicalKey(r.expt)===canonicalKey(state.expt))&&(state.type==='all'||r.vehicleType===state.type)&&(state.fleet==='all'||r.fleetType===state.fleet)}
function filteredDetailRows(periods=currentPeriods()){return DATA.unifiedRows.filter(r=>passFilters(r,periods))}
function filteredBudgetRows(periods=currentPeriods()){return (DATA.budgetRows||[]).filter(r=>passFilters(r,periods))}
function filteredOpRows(periods=currentPeriods()){return (DATA.operationRows||[]).filter(r=>passFilters(r,periods))}
function finalizeBucket(bucket){const gross=bucket.gross||0,deductions=(bucket.taxes||0)+(bucket.deductions||0),netRevenue=gross+deductions,costsVariable=bucket.costsVariable||0,costsTotal=(bucket.costs||0)+(bucket.otherRevenue||0),marginBeforeFixedCosts=netRevenue+costsVariable,costsFixed=costsTotal-costsVariable,contributionMargin=netRevenue+costsTotal,expensesTotal=(bucket.admin||0)+(bucket.people||0)+(bucket.sales||0)+(bucket.otherTaxes||0),ebitda=contributionMargin+expensesTotal,depreciation=bucket.depreciation||0,ebit=ebitda+depreciation,financialResult=(bucket.financialRevenue||0)+(bucket.financialExpense||0),ebt=ebit+financialResult,resultTaxes=(bucket.irpj||0)+(bucket.irrf||0)+(bucket.csll||0),netResult=ebt+resultTaxes;return {...bucket,gross,deductions,netRevenue,costsVariable,marginBeforeFixedCosts,costsFixed,costsTotal,contributionMargin,expensesTotal,ebitda,depreciation,ebit,financialResult,ebt,resultTaxes,netResult}}
function isVariableCost(r){const t=normText(r.costType);if(t.includes('variavel'))return true;if(t.includes('fixo'))return false;const c=normText(r.category);return c.includes('servicos agregados')}
function financeFromDetails(rows){const bucket={};for(const r of rows){const field=detailField(r);bucket[field]=(bucket[field]||0)+(r.value||0);if(field==='costs'&&isVariableCost(r))bucket.costsVariable=(bucket.costsVariable||0)+(r.value||0)}return finalizeBucket(bucket)}
function scopedFinance(periods){return [financeFromDetails(filteredDetailRows(periods))]}
function scopedBudget(periods){return [financeFromDetails(filteredBudgetRows(periods))]}
function rowsFinance(){return scopedFinance(currentPeriods())}
function rowsOps(){return filteredOpRows(currentPeriods()).filter(isOperationalRow)}
function sumRows(rows){return rows.reduce((a,r)=>{for(const [k,v] of Object.entries(r))if(typeof v==='number')a[k]=(a[k]||0)+v;return a},{})}
function weightedOps(rows){const s=sumRows(rows);s.performance=s.performanceWeight?s.performanceWeighted/s.performanceWeight:0;s.evidenceRate=s.routes?s.evidenced/s.routes:0;return s}
function renderKpis(target='kpis'){const f=sumRows(rowsFinance());const o=weightedOps(rowsOps());const margin=f.netRevenue?f.contributionMargin/f.netRevenue:0;const ebitda=f.netRevenue?f.ebitda/f.netRevenue:0;$(target).innerHTML=[
['Faturamento Líquido',fmtMoney(f.netRevenue),'Valores em R$ mil','var(--blue)',f.netRevenue],['Margem de Contribuição',fmtMoney(f.contributionMargin),fmtPct(margin)+' da RL','var(--green)',f.contributionMargin],['Custos / Despesas',fmtMoney(Math.abs(f.costsTotal||0)+Math.abs(f.expensesTotal||0)),'pressão operacional','var(--red)',-(Math.abs(f.costsTotal||0)+Math.abs(f.expensesTotal||0))],['EBITDA',fmtMoney(f.ebitda),fmtPct(ebitda)+' da RL','var(--teal)',f.ebitda],['Performance Entrega',fmtPct(o.performance),`${(o.delivered||0).toLocaleString('pt-BR')} entregues`,'var(--violet)',o.performance]
].map(c=>`<article class="card kpi" style="--accent:${c[3]}"><button type="button" class="note-btn" data-note="${target}:kpi:${c[0]}" data-note-title="${c[0]}" title="Notas explicativas" aria-label="Editar nota explicativa">${noteIconSvg()}</button><div class="kpi-label">${c[0]}</div><div class="kpi-value ${cls(c[4])}">${c[1]}</div><div class="kpi-note">${c[2]}</div></article>`).join('')}
function monthlyData(){const periods=currentPeriods();return periods.map(p=>{const f=financeFromDetails(filteredDetailRows([p]));const o=weightedOps(filteredOpRows([p]));return {period:p,...f,routes:o.routes||0,performanceWeighted:o.performanceWeighted||0,performanceWeight:o.performanceWeight||0}}).filter(d=>Math.abs(d.netRevenue)>0||d.routes>0)}
function renderMonthly(){const data=monthlyData();const w=760,h=300,pad=32,bottom=42,top=30,plot=h-top-bottom;const max=Math.max(...data.flatMap(d=>[Math.abs(d.netRevenue),Math.abs(d.ebitda),Math.abs(d.netResult)]),1);const step=(w-pad*2)/Math.max(data.length,1),bar=Math.min(34,step*.42);let bars='',eb=[],res=[];data.forEach((d,i)=>{const x=pad+i*step+step/2;const z=h-bottom;const bh=Math.abs(d.netRevenue)/max*(plot-34);const ye=z-d.ebitda/max*(plot-34),yr=z-d.netResult/max*(plot-34);eb.push(`${x},${ye}`);res.push(`${x},${yr}`);bars+=`<rect x="${x-bar/2}" y="${z-bh}" width="${bar}" height="${bh}" rx="6" fill="#0d5be1"></rect><text x="${x}" y="${Math.max(14,z-bh-9)}" text-anchor="middle" font-size="9" font-weight="850">${fmtMoney(d.netRevenue)}</text><text x="${x}" y="${h-18}" text-anchor="middle" font-size="10" font-weight="750" fill="#63708d">${shortMonth(d.period)}</text><circle cx="${x}" cy="${ye}" r="4" fill="#0aa36f"></circle><circle cx="${x}" cy="${yr}" r="4" fill="#7c3aed"></circle>`});$('monthlyChart').innerHTML=`<svg viewBox="0 0 ${w} ${h}" class="chart"><line x1="${pad}" x2="${w-pad}" y1="${h-bottom}" y2="${h-bottom}" stroke="#dce5f3"/><line x1="${pad}" x2="${w-pad}" y1="${top}" y2="${top}" stroke="#dce5f3"/><text x="${pad}" y="${top-9}" font-size="9" font-weight="750" fill="#63708d">Valores em R$ mil</text>${bars}<polyline points="${eb.join(' ')}" fill="none" stroke="#0aa36f" stroke-width="4"/><polyline points="${res.join(' ')}" fill="none" stroke="#7c3aed" stroke-width="4"/></svg><div class="legend"><span><i class="sw" style="background:#0d5be1"></i>Receita líquida</span><span><i class="sw" style="background:#0aa36f"></i>EBITDA</span><span><i class="sw" style="background:#7c3aed"></i>Resultado líquido</span></div>`}
function normText(s){return String(s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toLowerCase()}
function detailField(r){const a=normText(r.account),c=normText(r.category);if(a.includes('rendimentos de aplicacoes')||c.includes('rendimentos de aplicacoes'))return'financialRevenue';if(a.includes('depreciacao')||c.includes('depreciacao'))return'depreciation';if(a.includes('irrf')||c.includes('irrf'))return'irrf';if(a.includes('irpj')||c.includes('irpj'))return'irpj';if(a.includes('csll')||c.includes('csll'))return'csll';if(a.includes('receita bruta'))return'gross';if(a.includes('impostos'))return'taxes';if(a.includes('deducoes'))return'deductions';if(a.includes('custo dos servicos'))return'costs';if(a.includes('outras receitas'))return'otherRevenue';if(a.includes('despesas administrativas'))return'admin';if(a.includes('despesas com pessoal'))return'people';if(a.includes('vendas e marketing'))return'sales';if(a.includes('outros tributos'))return'otherTaxes';if(a.includes('receitas financeiras'))return'financialRevenue';if(a.includes('despesas financeiras'))return'financialExpense';return'other'}
function lineBucketForRow(r){const field=detailField(r);if(field==='taxes'||field==='deductions')return'deductions';if(field==='costs')return isVariableCost(r)?'costsVariable':'costsFixed';if(field==='otherRevenue')return'costsFixed';if(['admin','people','sales','otherTaxes'].includes(field))return'expensesTotal';if(field==='depreciation')return'depreciation';if(field==='financialRevenue'||field==='financialExpense')return'financialResult';if(field==='irpj'||field==='irrf'||field==='csll')return'resultTaxes';return field}
function lineBucketForField(field){if(field==='taxes'||field==='deductions')return'deductions';if(field==='costs')return'costsFixed';if(field==='otherRevenue')return'costsFixed';if(['admin','people','sales','otherTaxes'].includes(field))return'expensesTotal';if(field==='depreciation')return'depreciation';if(field==='financialRevenue'||field==='financialExpense')return'financialResult';if(field==='irpj'||field==='irrf'||field==='csll')return'resultTaxes';return field}
function detailRows(periods,lineKey){return filteredDetailRows(periods).filter(r=>lineBucketForRow(r)===lineKey)}
function aggregateDetailsFromRows(rows,lineKey){const accounts=new Map();for(const r of rows.filter(r=>lineBucketForRow(r)===lineKey)){const account=r.account||'N/D';const category=r.category||'N/D';if(!accounts.has(account))accounts.set(account,{label:account,periods:{},total:0,categories:new Map()});const acc=accounts.get(account);acc.periods[r.period]=(acc.periods[r.period]||0)+(r.value||0);acc.total+=(r.value||0);if(!acc.categories.has(category))acc.categories.set(category,{label:category,periods:{},total:0});const cat=acc.categories.get(category);cat.periods[r.period]=(cat.periods[r.period]||0)+(r.value||0);cat.total+=(r.value||0)}return [...accounts.values()].filter(d=>Math.abs(d.total)>.0001).map(d=>({...d,categories:[...d.categories.values()].filter(c=>Math.abs(c.total)>.0001).sort((a,b)=>Math.abs(b.total)-Math.abs(a.total))})).sort((a,b)=>Math.abs(b.total)-Math.abs(a.total))}
function aggregateDetails(periods,lineKey){return aggregateDetailsFromRows(detailRows(periods,lineKey),lineKey)}
function registerLedger(query,title){const key='ledger:'+LEDGER_CONTEXTS.size;LEDGER_CONTEXTS.set(key,{query,title});return key}
function bindLedgerRows(scope=document){scope.querySelectorAll('[data-ledger-key]').forEach(row=>{row.ondblclick=e=>{const key=row.dataset.ledgerKey,ctx=LEDGER_CONTEXTS.get(key);if(!ctx)return;openLedgerPopup(ctx.query,ctx.title)}})}
function commonText(rows,fn,label){const values=[...new Set(rows.map(fn).filter(Boolean))];return values.length===1?`${label}: ${values[0]}`:values.length?`${label}: Diversos`:''}
async function openLedgerPopup(query,title){$('ledgerTitle').textContent=`Lançamentos - ${title}`;$('ledgerSubtitle').textContent='Carregando lançamentos...';$('ledgerTable').innerHTML='';$('ledgerModal').hidden=false;const payload={...query,filters:{year:state.year,months:currentPeriods(),client:state.client,project:state.project,unit:state.unit,expt:state.expt,type:state.type,fleet:state.fleet}};try{const res=await fetch(new URL('/ledger',location.origin).href,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload),cache:'no-store'});const data=await res.json();if(!res.ok||!data.ok)throw new Error(data.error||'Não foi possível carregar os lançamentos.');const rows=data.rows||[],absorbed=data.absorbed||0,original=data.original||0,context=data.context||'',count=data.count||rows.length;$('ledgerSubtitle').textContent=`${count.toLocaleString('pt-BR')} lançamentos | Valor total ${fmtMoney(original)} | Valor absorvido ${fmtMoney(absorbed)}${context?' | '+context:''}`;$('ledgerTable').innerHTML=`<thead><tr><th>Período</th><th>Origem</th><th>Fornecedor/Cliente</th><th>NF</th><th>Valor Total</th><th>Valor Rateado / Absorvido</th></tr></thead><tbody>${rows.map(r=>{const total=Number.isFinite(Number(r.originalValue))?Number(r.originalValue):r.value||0;return `<tr><td>${esc(monthLabel(r.period))}</td><td>${esc(r.source||'')}</td><td>${esc(displayLabel(r.party||r.client||''))}</td><td>${esc(r.invoice||'')}</td><td class="${cls(total)}">${fmtMoney(total)}</td><td class="${cls(r.value)}">${fmtMoney(r.value)}</td></tr>`}).join('')}</tbody>`}catch(err){$('ledgerSubtitle').textContent='Erro ao carregar lançamentos: '+err.message}}
function closeLedgerPopup(){$('ledgerModal').hidden=true}
function plLines(f){return [['R$ Receita Bruta','gross',f.gross,true,false],['(-) R$ Deduções / Impostos','deductions',f.deductions,false,true],['(=) R$ Receita Líquida','netRevenue',f.netRevenue,true,false],['(-) R$ Custos Variáveis','costsVariable',f.costsVariable,false,true],['(=) R$ Margem de Contribuição antes dos Custos Fixos','marginBeforeFixedCosts',f.marginBeforeFixedCosts,true,false],['(-) R$ Custos Fixos','costsFixed',f.costsFixed,false,true],['(=) R$ Margem de Contribuição','contributionMargin',f.contributionMargin,true,false],['(-) R$ Despesas','expensesTotal',f.expensesTotal,false,true],['(=) R$ EBITDA','ebitda',f.ebitda,true,false],['(-) R$ Depreciações','depreciation',f.depreciation,false,true],['(=) R$ EBIT','ebit',f.ebit,true,false],['(-) Resultado Financeiro','financialResult',f.financialResult,false,true],['(=) R$ EBT','ebt',f.ebt,true,false],['(-) R$ Impostos antes do Resultado','resultTaxes',f.resultTaxes,false,true],['(=) R$ Resultado Líquido','netResult',f.netResult,true,false]]}
function renderTable(target='plTable'){const f=sumRows(rowsFinance());const lines=plLines(f);$(target).innerHTML=`<thead><tr><th>Linha P&L</th><th>Valor</th><th>% RL</th></tr></thead><tbody>${lines.map(l=>`<tr class="${l[3]?'total-row':''}"><td>${l[0]}</td><td class="${cls(l[2])}">${fmtMoney(l[2])}</td><td>${f.netRevenue?fmtPct(Math.abs(l[2])/f.netRevenue):'-'}</td></tr>`).join('')}</tbody>`}
function renderBudgetKpis(){const periods=currentPeriods(),actual=sumRows(scopedFinance(periods)),budget=sumRows(scopedBudget(periods));const variance=(key)=>(actual[key]||0)-(budget[key]||0),pct=(key)=>variationPct(actual[key]||0,budget[key]||0);const label=periodRangeLabel(periods);const cards=[
['Receita Líquida Real.',fmtMoney(actual.netRevenue),'Realizado','var(--blue)',actual.netRevenue],
['Receita Líquida Budget',fmtMoney(budget.netRevenue),`Budget ${label}`,'var(--green)',budget.netRevenue],
['Variação Receita',fmtMoney(variance('netRevenue')),fmtPct(pct('netRevenue')),'var(--red)',variance('netRevenue')],
['EBITDA vs Budget',fmtMoney(variance('ebitda')),fmtPct(pct('ebitda')),'var(--teal)',variance('ebitda')],
['Resultado vs Budget',fmtMoney(variance('netResult')),fmtPct(pct('netResult')),'var(--violet)',variance('netResult')]
];$('budgetKpis').innerHTML=cards.map(c=>`<article class="card kpi" style="--accent:${c[3]}"><button type="button" class="note-btn" data-note="budget:kpi:${c[0]}" data-note-title="${c[0]}" title="Notas explicativas" aria-label="Editar nota explicativa">${noteIconSvg()}</button><div class="kpi-label">${c[0]}</div><div class="kpi-value ${cls(c[4])}">${c[1]}</div><div class="kpi-note">${c[2]}</div></article>`).join('')}
function budgetMonthlyData(){return currentPeriods().map(p=>{const actual=financeFromDetails(filteredDetailRows([p])),budget=financeFromDetails(filteredBudgetRows([p]));return {period:p,actual,budget,diff:(actual.netRevenue||0)-(budget.netRevenue||0)}}).filter(d=>Math.abs(d.actual.netRevenue||0)>0||Math.abs(d.budget.netRevenue||0)>0)}
function renderBudgetMonthly(){const data=budgetMonthlyData();const w=760,h=260,pad=34,bottom=44,top=28,plot=h-top-bottom;const max=Math.max(...data.flatMap(d=>[Math.abs(d.actual.netRevenue||0),Math.abs(d.budget.netRevenue||0),Math.abs(d.actual.ebitda||0),Math.abs(d.budget.ebitda||0),Math.abs(d.budget.netResult||0)]),1);const step=(w-pad*2)/Math.max(data.length,1),bar=Math.min(18,step*.24);let bars='',ebitdaActual=[],ebitdaBudget=[],resultBudget=[];const point=(value)=>h-bottom-value/max*(plot-32);data.forEach((d,i)=>{const x=pad+i*step+step/2,z=h-bottom,actual=d.actual.netRevenue||0,budget=d.budget.netRevenue||0,bha=Math.abs(actual)/max*(plot-32),bhb=Math.abs(budget)/max*(plot-32);if(Math.abs(actual)>0||Math.abs(d.actual.ebitda||0)>0)ebitdaActual.push(`${x},${point(d.actual.ebitda||0)}`);if(Math.abs(budget)>0||Math.abs(d.budget.ebitda||0)>0)ebitdaBudget.push(`${x},${point(d.budget.ebitda||0)}`);if(Math.abs(budget)>0||Math.abs(d.budget.netResult||0)>0)resultBudget.push(`${x},${point(d.budget.netResult||0)}`);bars+=`<rect x="${x-bar-2}" y="${z-bha}" width="${bar}" height="${bha}" rx="5" fill="#0d5be1"><title>${shortMonth(d.period)} Realizado RL: ${fmtMoney(actual)}</title></rect><rect x="${x+2}" y="${z-bhb}" width="${bar}" height="${bhb}" rx="5" fill="#0aa36f" opacity=".75"><title>${shortMonth(d.period)} Budget RL: ${fmtMoney(budget)}</title></rect><text x="${x}" y="${h-18}" text-anchor="middle" font-size="10" font-weight="750" fill="#63708d">${shortMonth(d.period)}</text>`});$('budgetMonthlyChart').innerHTML=`<svg viewBox="0 0 ${w} ${h}" class="chart"><line x1="${pad}" x2="${w-pad}" y1="${h-bottom}" y2="${h-bottom}" stroke="#dce5f3"/><line x1="${pad}" x2="${w-pad}" y1="${top}" y2="${top}" stroke="#dce5f3"/><text x="${pad}" y="${top-9}" font-size="9" font-weight="750" fill="#63708d">Valores em R$ mil</text>${bars}<polyline points="${ebitdaActual.join(' ')}" fill="none" stroke="#7c3aed" stroke-width="4"><title>EBITDA realizado</title></polyline><polyline points="${ebitdaBudget.join(' ')}" fill="none" stroke="#0891b2" stroke-width="4" stroke-dasharray="7 5"><title>EBITDA budget</title></polyline><polyline points="${resultBudget.join(' ')}" fill="none" stroke="#f59e0b" stroke-width="4" stroke-dasharray="4 5"><title>Resultado líquido budget</title></polyline></svg><div class="legend"><span><i class="sw" style="background:#0d5be1"></i>Realizado RL</span><span><i class="sw" style="background:#0aa36f"></i>Budget RL</span><span><i class="sw" style="background:#7c3aed"></i>EBITDA realizado</span><span><i class="sw" style="background:#0891b2"></i>EBITDA budget</span><span><i class="sw" style="background:#f59e0b"></i>Resultado budget</span></div>`}
function renderBudgetTable(){const periods=currentPeriods(),actual=sumRows(scopedFinance(periods)),budget=sumRows(scopedBudget(periods)),rows=plLines(actual);const cell=(value,net)=>`${fmtMoney(value)}<span class="sub">% RL ${net?fmtPct(Math.abs(value)/net):'-'}</span>`;$('budgetTable').innerHTML=`<thead><tr><th>Linha P&L</th><th>Realizado</th><th>Budget</th><th>Variação R$</th><th>Variação %</th></tr></thead><tbody>${rows.map(line=>{const key=line[1],a=actual[key]||0,b=budget[key]||0,d=a-b,p=variationPct(a,b);return `<tr class="${line[3]?'total-row':line[4]?'group-row':''}"><td>${line[0]}</td><td class="${cls(a)}">${cell(a,actual.netRevenue)}</td><td class="${cls(b)}">${cell(b,budget.netRevenue)}</td><td class="${cls(d)}">${fmtMoney(d)}</td><td class="${Number.isFinite(p)?cls(p):''}">${Number.isFinite(p)?fmtPct(p):'-'}</td></tr>`}).join('')}</tbody>`}
function renderBudgetByClient(){const periods=currentPeriods();const groups=new Map();for(const r of filteredBudgetRows(periods)){if(!groups.has(r.client))groups.set(r.client,[]);groups.get(r.client).push(r)}const actualGroups=new Map();for(const r of filteredDetailRows(periods)){if(!actualGroups.has(r.client))actualGroups.set(r.client,[]);actualGroups.get(r.client).push(r)}const names=[...new Set([...groups.keys(),...actualGroups.keys()])].filter(Boolean);const data=names.map(name=>{const a=financeFromDetails(actualGroups.get(name)||[]),b=financeFromDetails(groups.get(name)||[]);return {name,a:a.netRevenue||0,b:b.netRevenue||0,d:(a.netRevenue||0)-(b.netRevenue||0)}}).filter(x=>Math.abs(x.a)>0||Math.abs(x.b)>0).sort((x,y)=>Math.abs(y.d)-Math.abs(x.d)).slice(0,12);$('budgetClientTable').innerHTML=`<thead><tr><th>Cliente</th><th>Realizado RL</th><th>Budget RL</th><th>Variação</th></tr></thead><tbody>${data.map(r=>`<tr><td>${displayLabel(r.name)}</td><td class="${cls(r.a)}">${fmtMoney(r.a)}</td><td class="${cls(r.b)}">${fmtMoney(r.b)}</td><td class="${cls(r.d)}">${fmtMoney(r.d)}</td></tr>`).join('')}</tbody>`}
function renderBudgetView(){const status=$('budgetRefreshStatus');if(status&&!status.textContent){const meta=DATA.meta||{};const parts=[];if(meta.budgetCampaigns!==undefined)parts.push(`${meta.budgetCampaigns} campanhas lidas`);if(meta.budgetRefreshedAt)parts.push(`atualizado em ${meta.budgetRefreshedAt}`);status.textContent=parts.join(' | ')}renderBudgetKpis();renderBudgetMonthly();renderBudgetTable();renderBudgetByClient()}
function previousPeriod(period){const [year,month]=period.split('-').map(Number);const d=new Date(year,month-2,1);return `${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,'0')}`}
function previousYearPeriod(period){const [year,month]=period.split('-').map(Number);return `${year-1}-${String(month).padStart(2,'0')}`}
function periodRangeLabel(periods){const list=[...periods].sort();if(!list.length)return'-';return list.length===1?monthLabel(list[0]):`${monthLabel(list[0])} a ${monthLabel(list[list.length-1])}`}
function variationPct(current,previous){return Number.isFinite(previous)&&Math.abs(previous)>0.01?(current-previous)/Math.abs(previous):null}
function renderVs(current,previous){const v=variationPct(current,previous);return Number.isFinite(v)?`<span class="sub ${v>=0?'value-pos':'value-neg'}">vs ${fmtPct(v)}</span>`:`<span class="sub">vs -</span>`}
function renderPlPeriodComparison(){const periods=currentPeriods();const previous=periods.map(previousYearPeriod).filter(p=>DATA.meta.periods.includes(p));const current=sumRows(scopedFinance(periods));const prior=sumRows(scopedFinance(previous));const rows=plLines(current);$('plComparisonNote').textContent=`${periodRangeLabel(periods)} vs ${periodRangeLabel(previous)}`;const cell=(value,net)=>`${fmtMoney(value)}<span class="sub">% RL ${net?fmtPct(Math.abs(value)/net):'-'}</span>`;$('plComparisonTable').innerHTML=`<thead><tr><th>Linha P&L</th><th>Período atual</th><th>Período comp.</th><th>Variação</th><th>Variação %</th></tr></thead><tbody>${rows.map(line=>{const key=line[1],cur=current[key]||0,old=prior[key]||0,diff=cur-old,pct=variationPct(cur,old);return `<tr class="${line[3]?'total-row':line[4]?'group-row':''}"><td>${line[0]}</td><td class="${cls(cur)}">${cell(cur,current.netRevenue)}</td><td class="${cls(old)}">${previous.length?cell(old,prior.netRevenue):'-'}</td><td class="${cls(diff)}">${previous.length?fmtMoney(diff):'-'}</td><td class="${Number.isFinite(pct)?(pct>=0?'value-pos':'value-neg'):''}">${Number.isFinite(pct)?fmtPct(pct):'-'}</td></tr>`}).join('')}</tbody>`}
function renderPlModuleTable(){const periods=currentPeriods();const buckets=new Map(periods.map(p=>[p,sumRows(scopedFinance([p]))]));const total=sumRows(scopedFinance(periods));const defs=plLines(total);const cell=(period,key)=>{const bucket=buckets.get(period)||{};const value=bucket[key]||0;const share=bucket.netRevenue?Math.abs(value)/bucket.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${fmtPct(share)}</span></td>`};const totalCell=(key)=>{const value=total[key]||0;const share=total.netRevenue?Math.abs(value)/total.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${fmtPct(share)}</span></td>`};const detailCell=(period,periodMap,net)=>{const value=periodMap[period]||0;const share=net?Math.abs(value)/net:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${share==null?'-':fmtPct(share)}</span></td>`};const detailTotalCell=(periodMap)=>{const value=periods.reduce((sum,p)=>sum+(periodMap[p]||0),0);const share=total.netRevenue?Math.abs(value)/total.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${share==null?'-':fmtPct(share)}</span></td>`};let body='';for(const line of defs){const open=!!state.drill[line[1]];const button=line[4]?`<button class="drill-btn" data-drill="${line[1]}">${open?'−':'+'}</button>`:'';body+=`<tr class="${line[3]?'total-row':line[4]?'group-row':''}"><td>${button}${line[0]}</td>${periods.map(p=>cell(p,line[1])).join('')}${totalCell(line[1])}</tr>`;if(open){for(const acc of aggregateDetails(periods,line[1])){const ledgerKey=registerLedger({lineKey:line[1],account:acc.label,category:null},acc.label);const accKey=line[1]+'|'+acc.label;const accOpen=!!state.drill[accKey];const accButton=acc.categories.length?`<button class="drill-btn" data-drill="${accKey}">${accOpen?'−':'+'}</button>`:'';body+=`<tr class="detail-row ledger-dbl" data-ledger-key="${ledgerKey}" title="Duplo clique para ver os lançamentos"><td class="detail-label">${accButton}${acc.label}</td>${periods.map(p=>detailCell(p,acc.periods,(buckets.get(p)||{}).netRevenue)).join('')}${detailTotalCell(acc.periods)}</tr>`;if(accOpen){for(const cat of acc.categories){const catLedgerKey=registerLedger({lineKey:line[1],account:acc.label,category:cat.label},cat.label);body+=`<tr class="detail-row ledger-dbl" data-ledger-key="${catLedgerKey}" title="Duplo clique para ver os lançamentos"><td class="detail-label" style="padding-left:52px">${cat.label}</td>${periods.map(p=>detailCell(p,cat.periods,(buckets.get(p)||{}).netRevenue)).join('')}${detailTotalCell(cat.periods)}</tr>`}}}}}$('plModuleTable').innerHTML=`<thead><tr><th>Linha P&L</th>${periods.map(p=>`<th>${monthLabel(p)}</th>`).join('')}<th>Total</th></tr></thead><tbody>${body}</tbody>`;document.querySelectorAll('[data-drill]').forEach(btn=>btn.onclick=()=>{state.drill[btn.dataset.drill]=!state.drill[btn.dataset.drill];renderPlModuleTable()});bindLedgerRows($('plModuleTable'))}
function renderUnitPlTable(){const rows=filteredDetailRows(currentPeriods());const groups=new Map();for(const r of rows){const key=hubExptLabel(r);if(!key||key==='N/D')continue;if(!groups.has(key))groups.set(key,[]);groups.get(key).push(r)}const groupList=[...groups.entries()].map(([name,items])=>({name,rows:items,finance:financeFromDetails(items)})).filter(g=>Math.abs(g.finance.netRevenue||0)>0||Math.abs(g.finance.netResult||0)>0||Math.abs(g.finance.ebitda||0)>0).sort((a,b)=>Math.abs(b.finance.netRevenue||0)-Math.abs(a.finance.netRevenue||0));const total=financeFromDetails(rows);const defs=plLines(total);const cell=(finance,key)=>{const value=finance[key]||0;const share=finance.netRevenue?Math.abs(value)/finance.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${fmtPct(share)}</span></td>`};const detailValue=(items,lineKey,account,category)=>items.filter(r=>lineBucketForRow(r)===lineKey&&(!account||r.account===account)&&(!category||r.category===category)).reduce((s,r)=>s+(r.value||0),0);const detailCell=(items,lineKey,account,category)=>{const finance=financeFromDetails(items);const value=detailValue(items,lineKey,account,category);const share=finance.netRevenue?Math.abs(value)/finance.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${share==null?'-':fmtPct(share)}</span></td>`};const totalDetailCell=(lineKey,account,category)=>{const value=detailValue(rows,lineKey,account,category);const share=total.netRevenue?Math.abs(value)/total.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${share==null?'-':fmtPct(share)}</span></td>`};let body='';for(const line of defs){const open=!!state.drill['unit:'+line[1]];const button=line[4]?`<button class="drill-btn" data-drill="unit:${line[1]}">${open?'−':'+'}</button>`:'';body+=`<tr class="${line[3]?'total-row':line[4]?'group-row':''}"><td>${button}${line[0]}</td>${groupList.map(g=>cell(g.finance,line[1])).join('')}${cell(total,line[1])}</tr>`;if(open){for(const acc of aggregateDetailsFromRows(rows,line[1])){const ledgerKey=registerLedger({lineKey:line[1],account:acc.label,category:null},acc.label);const accKey='unit:'+line[1]+'|'+acc.label;const accOpen=!!state.drill[accKey];const accButton=acc.categories.length?`<button class="drill-btn" data-drill="${accKey}">${accOpen?'−':'+'}</button>`:'';body+=`<tr class="detail-row ledger-dbl" data-ledger-key="${ledgerKey}" title="Duplo clique para ver os lançamentos"><td class="detail-label">${accButton}${acc.label}</td>${groupList.map(g=>detailCell(g.rows,line[1],acc.label,null)).join('')}${totalDetailCell(line[1],acc.label,null)}</tr>`;if(accOpen){for(const cat of acc.categories){const catLedgerKey=registerLedger({lineKey:line[1],account:acc.label,category:cat.label},cat.label);body+=`<tr class="detail-row ledger-dbl" data-ledger-key="${catLedgerKey}" title="Duplo clique para ver os lançamentos"><td class="detail-label" style="padding-left:52px">${cat.label}</td>${groupList.map(g=>detailCell(g.rows,line[1],acc.label,cat.label)).join('')}${totalDetailCell(line[1],acc.label,cat.label)}</tr>`}}}}}const empty='<tr><td colspan="2">Sem dados para o filtro selecionado.</td></tr>';$('unitPlTable').innerHTML=`<thead><tr><th>Linha P&L</th>${groupList.map(g=>`<th>${g.name}</th>`).join('')}<th>Total</th></tr></thead><tbody>${groupList.length?body:empty}</tbody>`;document.querySelectorAll('#unitPlTable [data-drill]').forEach(btn=>btn.onclick=()=>{state.drill[btn.dataset.drill]=!state.drill[btn.dataset.drill];renderUnitPlTable()});bindLedgerRows($('unitPlTable'))}
function rankBy(level,field,target,title){const groups=new Map();for(const r of filteredDetailRows(currentPeriods())){const key=level==='client'?r.client:level==='hub'?r.hub:r.project;if(!key)continue;if(!groups.has(key))groups.set(key,[]);groups.get(key).push(r)}const data=[...groups.entries()].map(([name,rows])=>({name,value:financeFromDetails(rows)[field]||0})).filter(d=>Math.abs(d.value)>0).sort((a,b)=>Math.abs(b.value)-Math.abs(a.value)).slice(0,8);const max=Math.max(...data.map(d=>Math.abs(d.value)),1);$(target).innerHTML=`<div class="rank">${data.map(d=>`<div class="rank-row"><span>${d.name}</span><strong class="${cls(d.value)}">${fmtMoney(d.value)}</strong><div class="track"><div class="fill" style="width:${Math.abs(d.value)/max*100}%;background:${d.value>=0?'#0aa36f':'#ef4444'}"></div></div></div>`).join('')}</div>`}
function fmtNum(v){return (Number(v)||0).toLocaleString('pt-BR',{maximumFractionDigits:0})}
function fmtTicket(v){return BRL.format(Number.isFinite(v)?v:0)}
function div(a,b){return b? a/b : 0}
function valueSet(rows,fn){const s=new Set();for(const r of rows){const v=fn(r);if(v&&v!=='N/D'&&v!=='Todos')s.add(v)}return s}
function setHasOrOpen(set,value){return !set.size||set.has(value)}
function popupOperationRows(rows){const periods=currentPeriods();const clients=valueSet(rows,r=>r.client),projects=valueSet(rows,r=>r.project),units=valueSet(rows,r=>canonicalKey(unitValue(r))),expts=valueSet(rows,r=>canonicalKey(r.expt)),types=valueSet(rows,r=>r.vehicleType),fleets=valueSet(rows,r=>r.fleetType);return filteredOpRows(periods).filter(r=>setHasOrOpen(clients,r.client)&&setHasOrOpen(projects,r.project)&&setHasOrOpen(units,canonicalKey(unitValue(r)))&&setHasOrOpen(expts,canonicalKey(r.expt))&&setHasOrOpen(types,r.vehicleType)&&setHasOrOpen(fleets,r.fleetType))}
function renderBreakEven(rows){const f=financeFromDetails(rows),ops=weightedOps(popupOperationRows(rows));const burden=Math.abs(f.deductions||0)+Math.abs(f.costsTotal||0)+Math.abs(f.expensesTotal||0)+Math.abs(f.depreciation||0)+Math.abs(f.financialResult||0)+Math.abs(f.resultTaxes||0);const routes=ops.routes||0,delivered=ops.delivered||0;const ticketRoute=routes?f.gross/routes:0,ticketDelivery=delivered?f.gross/delivered:0;const requiredRoutes=ticketRoute?burden/ticketRoute:0,requiredDeliveries=ticketDelivery?burden/ticketDelivery:0;const gapRoutes=routes-requiredRoutes,gapDeliveries=delivered-requiredDeliveries;const count=v=>Number.isFinite(v)?Math.round(v).toLocaleString('pt-BR'):'-';$('unitPlBreakEven').innerHTML=`<div class="breakeven-grid"><div class="be-card"><div class="be-label">Necessidade financeira</div><div class="be-value">${fmtMoney(burden)}</div><div class="be-note">impostos, custos e demais saídas</div></div><div class="be-card"><div class="be-label">Ticket médio rota</div><div class="be-value">${fmtTicket(ticketRoute)}</div><div class="be-note">receita bruta / rotas</div></div><div class="be-card"><div class="be-label">Ticket médio entrega</div><div class="be-value">${fmtTicket(ticketDelivery)}</div><div class="be-note">receita bruta / entregues</div></div><div class="be-card"><div class="be-label">Breakeven rotas</div><div class="be-value">${count(requiredRoutes)}</div><div class="be-note">realizadas ${count(routes)} | gap ${count(gapRoutes)}</div></div><div class="be-card"><div class="be-label">Breakeven entregas</div><div class="be-value">${count(requiredDeliveries)}</div><div class="be-note">entregues ${count(delivered)} | gap ${count(gapDeliveries)}</div></div></div>`}
function grossByFleet(fleet){return filteredDetailRows(currentPeriods()).filter(r=>detailField(r)==='gross'&&(fleet==='Todos'||r.fleetType===fleet)).reduce((s,r)=>s+(r.value||0),0)}
function renderOps(target='opsTable'){const rows=rowsOps();const o=weightedOps(rows);const totalRoutes=o.routes||0,aggRoutes=o.aggregatedRoutes||0,fleetRoutes=o.fleetRoutes||0,shipped=o.shipped||0,delivered=o.delivered||0;const agg=weightedOps(rows.filter(r=>r.fleetType==='Agregado'));const fleet=weightedOps(rows.filter(r=>r.fleetType==='Frota'));const gross=grossByFleet('Todos'),grossAgg=grossByFleet('Agregado'),grossFleet=grossByFleet('Frota');const rowsHtml=[
['section','Rotas','',''],['row','Rotas Agregado',fmtNum(aggRoutes),fmtPct(div(aggRoutes,totalRoutes))],['row','Rotas Frota',fmtNum(fleetRoutes),fmtPct(div(fleetRoutes,totalRoutes))],['total','Total Rotas',fmtNum(totalRoutes),'100,0%'],
['section','Performance','',''],['row','Embarcadas',fmtNum(shipped),''],['row','Entregues',fmtNum(delivered),''],['total','Performance %',fmtPct(div(delivered,shipped)),''],
['section','SPR','',''],['row','SPR Agregados',fmtNum(div(agg.shipped||0,aggRoutes)),'Embarcados / Rotas'],['row','SPR Frota',fmtNum(div(fleet.shipped||0,fleetRoutes)),'Embarcados / Rotas'],['total','SPR',fmtNum(div(shipped,totalRoutes)),'Embarcados / Total Rotas'],
['section','Ticket Médio','',''],['row','Ticket Médio Pacote',fmtTicket(div(gross,delivered)),'Receita Bruta / Entregues'],['row','Ticket Médio Rota',fmtTicket(div(gross,totalRoutes)),'Receita Bruta / Rotas'],['row','Ticket Médio Rota Agregado',fmtTicket(div(grossAgg,aggRoutes)),'Receita Bruta Agregado / Rotas'],['row','Ticket Médio Rota Frota',fmtTicket(div(grossFleet,fleetRoutes)),'Receita Bruta Frota / Rotas']
].map(r=>r[0]==='section'?`<tr class="ops-section"><td colspan="3">${r[1]}</td></tr>`:`<tr class="${r[0]==='total'?'ops-total':''}"><td>${r[1]}</td><td>${r[2]}</td><td>${r[3]}</td></tr>`).join('');$(target).innerHTML=`<thead><tr><th>Indicador</th><th>Valor</th><th>% / Critério</th></tr></thead><tbody>${rowsHtml}</tbody>`}
function hubExptLabel(row){const hub=unitValue(row)||row.hub||'N/D',expt=row.expt||'N/D';return canonicalKey(hub)===canonicalKey(expt)?displayLabel(hub):`${displayLabel(hub)} / ${displayLabel(expt)}`}
function hierarchyLabel(row){return `${displayLabel(row.client||'N/D')} > ${displayLabel(row.project||'N/D')} > ${hubExptLabel(row)}`}
function unitRowsByLabel(label){return filteredDetailRows(currentPeriods()).filter(r=>hierarchyLabel(r)===label)}
function uniqueLabel(rows,key){const values=[...new Set(rows.map(r=>r[key]).filter(v=>v&&v!=='N/D'))];return values.length===1?displayLabel(values[0]):'Diversos'}
function unitPopupTitle(label,rows){return label}
function renderPopupPlRows(rows,label){const periods=currentPeriods();const buckets=new Map(periods.map(p=>[p,financeFromDetails(rows.filter(r=>r.period===p))]));const total=financeFromDetails(rows);const defs=plLines(total);const cell=(finance,key)=>{const value=finance[key]||0;const share=finance.netRevenue?Math.abs(value)/finance.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${fmtPct(share)}</span></td>`};const detailValue=(items,lineKey,account,category)=>items.filter(r=>lineBucketForRow(r)===lineKey&&(!account||r.account===account)&&(!category||r.category===category)).reduce((s,r)=>s+(r.value||0),0);const detailCell=(period,lineKey,account,category)=>{const periodRows=rows.filter(r=>r.period===period);const finance=buckets.get(period)||{};const value=detailValue(periodRows,lineKey,account,category);const share=finance.netRevenue?Math.abs(value)/finance.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${share==null?'-':fmtPct(share)}</span></td>`};const totalDetailCell=(lineKey,account,category)=>{const value=detailValue(rows,lineKey,account,category);const share=total.netRevenue?Math.abs(value)/total.netRevenue:null;return `<td class="${cls(value)}">${fmtMoney(value)}<span class="sub">% RL ${share==null?'-':fmtPct(share)}</span></td>`};let body='';for(const line of defs){const drillKey=`popup:${label}:${line[1]}`;const open=!!state.drill[drillKey];const button=line[4]?`<button class="drill-btn" data-popup-drill="${drillKey}">${open?'−':'+'}</button>`:'';body+=`<tr class="${line[3]?'total-row':line[4]?'group-row':''}"><td>${button}${line[0]}</td>${periods.map(p=>cell(buckets.get(p)||{},line[1])).join('')}${cell(total,line[1])}</tr>`;if(open){for(const acc of aggregateDetailsFromRows(rows,line[1])){const accKey=`${drillKey}|${acc.label}`;const accOpen=!!state.drill[accKey];const accButton=acc.categories.length?`<button class="drill-btn" data-popup-drill="${accKey}">${accOpen?'−':'+'}</button>`:'';body+=`<tr class="detail-row"><td class="detail-label">${accButton}${acc.label}</td>${periods.map(p=>detailCell(p,line[1],acc.label,null)).join('')}${totalDetailCell(line[1],acc.label,null)}</tr>`;if(accOpen){for(const cat of acc.categories){body+=`<tr class="detail-row"><td class="detail-label" style="padding-left:52px">${cat.label}</td>${periods.map(p=>detailCell(p,line[1],acc.label,cat.label)).join('')}${totalDetailCell(line[1],acc.label,cat.label)}</tr>`}}}}}return `<thead><tr><th>Linha P&L</th>${periods.map(p=>`<th>${monthLabel(p)}</th>`).join('')}<th>Total</th></tr></thead><tbody>${body}</tbody>`}
function bindPopupDrills(label){document.querySelectorAll('#unitPlModalTable [data-popup-drill]').forEach(btn=>btn.onclick=()=>{state.drill[btn.dataset.popupDrill]=!state.drill[btn.dataset.popupDrill];const rows=unitRowsByLabel(label);$('unitPlModalTable').innerHTML=renderPopupPlRows(rows,label);bindPopupDrills(label)})}
function openUnitPlPopup(label){const rows=unitRowsByLabel(label);const title=unitPopupTitle(label,rows);$('unitPlModalTitle').textContent=`P&L - ${title}`;$('unitPlModalSubtitle').textContent=`${periodRangeLabel(currentPeriods())} | ${rows.length.toLocaleString('pt-BR')} lançamentos`;$('unitPlModalTable').innerHTML=renderPopupPlRows(rows,label);renderBreakEven(rows);bindPopupDrills(label);$('unitPlNote').dataset.note=`unit-popup:${title}`;$('unitPlNote').dataset.noteTitle=`P&L - ${title}`;updateNoteButtons();$('unitPlModal').hidden=false}
function bindGenericPopupDrills(key,title){document.querySelectorAll('#unitPlModalTable [data-popup-drill]').forEach(btn=>btn.onclick=()=>{state.drill[btn.dataset.popupDrill]=!state.drill[btn.dataset.popupDrill];const rows=POPUP_GROUPS.get(key)||[];$('unitPlModalTable').innerHTML=renderPopupPlRows(rows,key);bindGenericPopupDrills(key,title)})}
function openGenericPlPopup(key,title){const rows=POPUP_GROUPS.get(key)||[];$('unitPlModalTitle').textContent=`P&L - ${title}`;$('unitPlModalSubtitle').textContent=`${periodRangeLabel(currentPeriods())} | ${rows.length.toLocaleString('pt-BR')} lançamentos`;$('unitPlModalTable').innerHTML=renderPopupPlRows(rows,key);renderBreakEven(rows);bindGenericPopupDrills(key,title);$('unitPlNote').dataset.note=`unit-popup:${title}`;$('unitPlNote').dataset.noteTitle=`P&L - ${title}`;updateNoteButtons();$('unitPlModal').hidden=false}
function closeUnitPlPopup(){$('unitPlModal').hidden=true}
function sortHeader(target){const arrow=(state.rankSort[target]||'desc')==='desc'?'↓':'↑';return `<th class="sort-th" data-sort-target="${target}">Resultado Líquido <span class="sort-arrow">${arrow}</span></th>`}
function bindSortHeader(target,renderer){const th=document.querySelector(`#${target} [data-sort-target]`);if(th)th.onclick=e=>{e.stopPropagation();state.rankSort[target]=(state.rankSort[target]||'desc')==='desc'?'asc':'desc';renderer()}}
function renderUnitResults(target='unitRank'){const groups=new Map();for(const r of filteredDetailRows(currentPeriods()).filter(isOperationalRow)){const key=hierarchyLabel(r);if(!key||key.includes('N/D'))continue;if(!groups.has(key))groups.set(key,[]);groups.get(key).push(r)}let data=[...groups.entries()].map(([name,rows])=>{const f=financeFromDetails(rows);return {name,displayName:name,ebitda:f.ebitda||0,net:f.netResult||0}}).filter(d=>Math.abs(d.ebitda)>0.01||Math.abs(d.net)>0.01).sort((a,b)=>(state.rankSort[target]||'desc')==='desc'?b.net-a.net:a.net-b.net).slice(0,50);if(!data.length){$(target).innerHTML='<div class="panel-note" style="text-align:left">Sem dados para o período filtrado.</div>';return}$(target).innerHTML=`<div class="table-wrap"><table class="table"><thead><tr><th>Cliente > Operação > Unidade / Expt</th><th>EBITDA</th>${sortHeader(target)}</tr></thead><tbody>${data.map(d=>`<tr class="click-row" data-unit-pl="${d.name}"><td>${d.displayName}</td><td class="${cls(d.ebitda)}">${fmtMoney(d.ebitda)}</td><td class="${cls(d.net)}">${fmtMoney(d.net)}</td></tr>`).join('')}</tbody></table></div>`;bindSortHeader(target,()=>renderUnitResults(target));document.querySelectorAll(`#${target} [data-unit-pl]`).forEach(row=>row.onclick=()=>openUnitPlPopup(row.dataset.unitPl))}
function renderDimensionResults(target,dimension){const groups=new Map();const dimLabel=dimension==='type'?'Tipo de Veículo':'Frota';for(const r of filteredDetailRows(currentPeriods()).filter(isOperationalRow)){const value=dimension==='type'?r.vehicleType:r.fleetType;if(!value||value==='N/D')continue;const label=`${hierarchyLabel(r)} > ${displayLabel(value)}`;if(!groups.has(label))groups.set(label,[]);groups.get(label).push(r)}let data=[...groups.entries()].map(([name,rows])=>{const f=financeFromDetails(rows);return {name,rows,ebitda:f.ebitda||0,net:f.netResult||0}}).filter(d=>Math.abs(d.ebitda)>0.01||Math.abs(d.net)>0.01).sort((a,b)=>(state.rankSort[target]||'desc')==='desc'?b.net-a.net:a.net-b.net).slice(0,50);if(!data.length){$(target).innerHTML='<div class="panel-note" style="text-align:left">Sem dados para o período filtrado.</div>';return}$(target).innerHTML=`<div class="table-wrap"><table class="table"><thead><tr><th>Cliente > Operação > Unidade / Expt > ${dimLabel}</th><th>EBITDA</th>${sortHeader(target)}</tr></thead><tbody>${data.map((d,i)=>{const key=`${target}:${i}`;POPUP_GROUPS.set(key,d.rows);return `<tr class="click-row" data-popup-key="${key}"><td>${d.name}</td><td class="${cls(d.ebitda)}">${fmtMoney(d.ebitda)}</td><td class="${cls(d.net)}">${fmtMoney(d.net)}</td></tr>`}).join('')}</tbody></table></div>`;bindSortHeader(target,()=>renderDimensionResults(target,dimension));document.querySelectorAll(`#${target} [data-popup-key]`).forEach(row=>row.onclick=()=>openGenericPlPopup(row.dataset.popupKey,row.cells[0].textContent))}
async function ensureBudgetData(){if(budgetDataLoaded)return;const payload=await (await fetch(BUDGET_DATA_URL+Date.now())).json();DATA.budgetRows=payload.budgetRows||[];DATA.meta={...(DATA.meta||{}),...(payload.meta||{})};budgetDataLoaded=true}
function render(){LEDGER_CONTEXTS.clear();document.body.classList.toggle('ready',true);refreshHierarchicalFilters();document.querySelectorAll('.view').forEach(v=>v.classList.toggle('active',v.id===`view-${state.view}`));document.querySelectorAll('[data-view]').forEach(b=>b.classList.toggle('active',b.dataset.view===state.view));$('pageTitle').textContent=state.view==='pl'?'P&L Gerencial':state.view==='budget'?'Budget x Realizado':state.view==='unitpl'?'P&L por Unidade / Expt':state.view==='ops'?'Operação':'Dashboard Executivo P&L';if(state.view==='dashboard'){renderKpis();renderMonthly();renderTable();renderOps();rankBy('client','netRevenue','clientRank');rankBy('hub','netRevenue','hubRank')}if(state.view==='pl'){renderKpis('plKpis');renderPlModuleTable()}if(state.view==='budget'){renderBudgetView()}if(state.view==='unitpl'){renderKpis('unitPlKpis');renderUnitPlTable()}if(state.view==='ops'){POPUP_GROUPS.clear();renderUnitResults('unitRank');renderDimensionResults('typeRank','type');renderDimensionResults('fleetRank','fleet');setupOpsCards()}ensureNoteButtons();updateNoteButtons()}
async function runRefresh(btn,url,{statusEl=null,successText='Atualizado'}={}){const original=btn.innerHTML;btn.disabled=true;btn.innerHTML='Atualizando...';if(statusEl)statusEl.textContent='Atualizando...';try{let response;try{response=await fetch(new URL(url,location.origin).href,{method:'POST',cache:'no-store'})}catch(firstErr){await new Promise(r=>setTimeout(r,1500));response=await fetch(new URL(url,location.origin).href,{method:'POST',cache:'no-store'})}const payload=await response.json().catch(()=>({}));if(!response.ok||!payload.ok){throw new Error(payload.stderr||payload.stdout||'Não foi possível atualizar a base.')}if(statusEl)statusEl.textContent=payload.stdout||successText;location.href='dashboard_pl.html?refresh='+Date.now()}catch(err){btn.disabled=false;btn.innerHTML=original;if(statusEl)statusEl.textContent='Falha ao atualizar';alert('Erro ao atualizar a base: '+err.message+'\\n\\nVerifique se o Excel não está salvando/sincronizando e tente novamente.')}}
async function init(){DATA=await (await fetch(DATA_URL)).json();await loadNotes();const ys=years();state.year=ys[0];fillSelect('yearFilter',ys);$('yearFilter').value=state.year;fillMonthChecks();refreshHierarchicalFilters();$('yearFilter').onchange=e=>{state.year=e.target.value;state.months=['all'];state.client=state.project=state.unit=state.expt=state.type=state.fleet='all';fillMonthChecks();render()};$('clientFilter').onchange=e=>{state.client=e.target.value;state.project=state.unit=state.expt=state.type=state.fleet='all';render()};$('projectFilter').onchange=e=>{state.project=e.target.value;state.unit=state.expt=state.type=state.fleet='all';render()};$('unitFilter').onchange=e=>{state.unit=e.target.value;state.expt=state.type=state.fleet='all';render()};$('exptFilter').onchange=e=>{state.expt=e.target.value;state.type=state.fleet='all';render()};$('typeFilter').onchange=e=>{state.type=e.target.value;state.fleet='all';render()};$('fleetFilter').onchange=e=>{state.fleet=e.target.value;render()};$('clearFilters').onclick=clearFilters;document.querySelectorAll('[data-view]').forEach(b=>b.onclick=async()=>{state.view=b.dataset.view;if(state.view==='budget')await ensureBudgetData();fillMonthChecks();render()});$('unitPlNote').innerHTML=noteIconSvg();document.addEventListener('click',e=>{const btn=e.target.closest('.note-btn');if(!btn)return;e.preventDefault();e.stopPropagation();openNote(btn.dataset.note,btn.dataset.noteTitle)});$('noteSave').onclick=saveCurrentNote;$('noteCancel').onclick=closeNote;$('noteClose').onclick=closeNote;$('noteModal').onclick=e=>{if(e.target===$('noteModal'))closeNote()};$('unitPlClose').onclick=closeUnitPlPopup;$('unitPlCloseBottom').onclick=closeUnitPlPopup;$('unitPlModal').onclick=e=>{if(e.target===$('unitPlModal'))closeUnitPlPopup()};$('ledgerClose').onclick=closeLedgerPopup;$('ledgerCloseBottom').onclick=closeLedgerPopup;$('ledgerModal').onclick=e=>{if(e.target===$('ledgerModal'))closeLedgerPopup()};$('refreshBase').onclick=()=>runRefresh($('refreshBase'),'/refresh');$('refreshBudget').onclick=()=>runRefresh($('refreshBudget'),'/refresh?view=budget',{statusEl:$('budgetRefreshStatus'),successText:'Budget atualizado'});render()}init();
"""


HTML = """<!doctype html>
<html lang="pt-BR">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Dashboard P&L Redefrete</title><style>{css}</style></head>
<body><div class="app"><aside class="sidebar"><div class="brand"><img class="brand-logo" src="assets/redefrete-logo-branco.png" alt="Redefrete"></div><nav class="nav"><button class="active" data-view="dashboard"><svg viewBox="0 0 24 24"><path d="M3 11l9-8 9 8"/><path d="M5 10v10h14V10"/></svg><span>Dashboard</span></button><button data-view="pl"><svg viewBox="0 0 24 24"><path d="M4 19h16"/><path d="M7 16v-5M12 16V7M17 16v-8"/></svg><span>P&L</span></button><button data-view="budget"><svg viewBox="0 0 24 24"><path d="M4 19h16"/><path d="M7 16V9M12 16V5M17 16v-4"/><path d="M4 5h4"/></svg><span>Budget</span></button><button data-view="unitpl"><svg viewBox="0 0 24 24"><path d="M4 5h16M4 12h16M4 19h16"/><path d="M8 5v14M16 5v14"/></svg><span>P&L Unidade/Expt</span></button><button data-view="ops"><svg viewBox="0 0 24 24"><path d="M4 17l6-6 4 4 6-8"/><path d="M15 7h5v5"/></svg><span>Operação</span></button></nav></aside><main class="main"><header class="topbar"><h1 id="pageTitle">Dashboard Executivo P&L</h1><div class="filter-card"><div class="control"><label>Ano</label><select id="yearFilter"></select></div><details class="control month-filter"><summary><span>Mês:</span><strong id="monthSummary">Todos</strong></summary><div class="check-list" id="monthChecks"></div></details><div class="control"><label>Cliente</label><select id="clientFilter"></select></div><div class="control"><label>Projeto</label><select id="projectFilter"></select></div><div class="control"><label>Unidade</label><select id="unitFilter"></select></div><div class="control"><label>Expt</label><select id="exptFilter"></select></div><div class="control"><label>Tipo</label><select id="typeFilter"></select></div><div class="control"><label>Frota</label><select id="fleetFilter"></select></div><button id="clearFilters">Limpar filtros</button><button class="refresh" id="refreshBase"><svg viewBox="0 0 24 24"><path d="M21 12a9 9 0 0 1-15.6 6.1"/><path d="M3 12A9 9 0 0 1 18.6 5.9"/><path d="M3 19v-5h5"/><path d="M21 5v5h-5"/></svg>Atualizar</button></div></header><section class="view active" id="view-dashboard"><section class="kpis" id="kpis"></section><section class="grid"><article class="panel wide" style="--accent:var(--green)"><h2>Evolução Mensal - Receita Líquida, EBITDA e Resultado Líquido</h2><div class="chart-wrap" id="monthlyChart"></div></article><article class="panel" style="--accent:var(--blue)"><h2>P&L Gerencial</h2><div class="table-wrap"><table class="table" id="plTable"></table></div></article><article class="panel" style="--accent:var(--teal)"><h2>Operação</h2><div class="table-wrap"><table class="table" id="opsTable"></table></div></article><article class="panel" style="--accent:var(--violet)"><h2>Receita Líquida por Cliente</h2><div id="clientRank"></div></article><article class="panel" style="--accent:var(--red)"><h2>Receita Líquida por HUB</h2><div id="hubRank"></div></article></section></section><section class="view" id="view-pl"><section class="kpis" id="plKpis"></section><div class="pl-module-grid"><article class="panel full" style="--accent:var(--blue)"><h2>P&L Detalhado</h2><div class="table-wrap pl-table-wrap"><table class="table" id="plModuleTable"></table></div></article></div></section><section class="view" id="view-budget"><div class="budget-actions"><span class="budget-status" id="budgetRefreshStatus"></span><button class="budget-refresh" id="refreshBudget" type="button">Atualizar budget</button></div><section class="kpis" id="budgetKpis"></section><section class="grid"><article class="panel wide" style="--accent:var(--green)"><h2>Budget x Realizado - Receita Líquida, EBITDA e Resultado Líquido</h2><div class="chart-wrap" id="budgetMonthlyChart"></div></article><article class="panel" style="--accent:var(--violet)"><h2>Variação por Cliente</h2><div class="table-wrap"><table class="table" id="budgetClientTable"></table></div></article><article class="panel full" style="--accent:var(--blue)"><h2>P&L Budget x Realizado</h2><div class="table-wrap pl-table-wrap"><table class="table" id="budgetTable"></table></div></article></section></section><section class="view" id="view-unitpl"><section class="kpis" id="unitPlKpis"></section><div class="pl-module-grid"><article class="panel full" style="--accent:var(--blue)"><h2>P&L por Unidade / Expt</h2><div class="table-wrap pl-table-wrap"><table class="table unit-pl-table" id="unitPlTable"></table></div></article></div></section><section class="view" id="view-ops"><div class="ops-rank-grid"><article class="panel" style="--accent:var(--green)"><h2>Resultado por Unidade / Expt</h2><div id="unitRank"></div></article><article class="panel" style="--accent:var(--teal)"><h2>Resultado por Tipo de Veículo</h2><div id="typeRank"></div></article><article class="panel" style="--accent:var(--violet)"><h2>Resultado por Frota</h2><div id="fleetRank"></div></article></div></section></main></div><div class="note-modal-backdrop" id="unitPlModal" hidden><section class="note-modal unit-modal"><div class="note-modal-head"><div><h3 id="unitPlModalTitle">P&L por Unidade / Expt</h3><p id="unitPlModalSubtitle"></p></div><button type="button" class="note-btn" id="unitPlNote" title="Notas explicativas" aria-label="Editar nota explicativa"></button><button class="note-close" id="unitPlClose" type="button">X</button></div><div id="unitPlBreakEven"></div><div class="unit-modal-body"><table class="table" id="unitPlModalTable"></table></div><div class="unit-modal-actions"><button class="secondary" id="unitPlCloseBottom" type="button">Fechar</button></div></section></div><div class="note-modal-backdrop" id="ledgerModal" hidden><section class="note-modal ledger-modal"><div class="note-modal-head"><div><h3 id="ledgerTitle">Lançamentos</h3><p id="ledgerSubtitle"></p></div><button class="note-close" id="ledgerClose" type="button">X</button></div><div class="ledger-modal-body"><table class="table ledger-table" id="ledgerTable"></table></div><div class="unit-modal-actions"><button class="secondary" id="ledgerCloseBottom" type="button">Fechar</button></div></section></div><div class="note-modal-backdrop" id="noteModal" hidden><section class="note-modal"><div class="note-modal-head"><div><h3 id="noteTitle">Notas explicativas</h3><p>As notas ficam salvas neste projeto e podem ser editadas a qualquer momento.</p></div><button class="note-close" id="noteClose" type="button">X</button></div><textarea id="noteText" placeholder="Escreva a explicação executiva deste card..."></textarea><div class="note-actions"><button class="secondary" id="noteCancel" type="button">Cancelar</button><button class="primary" id="noteSave" type="button">Salvar nota</button></div></section></div><script>{js}</script></body></html>"""


def write_html():
    HTML_OUT.write_text(HTML.format(css=CSS, js=JS), encoding="utf-8")


if __name__ == "__main__":
    data = build_data()
    write_html()
    print(HTML_OUT)
    print(DATA_OUT)
    print(f"finance aggregates: {len(data['finance'])}")
    print(f"operation aggregates: {len(data['operations'])}")







