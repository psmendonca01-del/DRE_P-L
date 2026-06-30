from __future__ import annotations

import importlib.util
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
WORKBOOK = BASE / "PL.xlsx"
LEDGER = BASE / "pl_ledger.json"
OUT = BASE / "auditoria_faturamento_pagamento.xlsx"


spec = importlib.util.spec_from_file_location("pl_builder", BASE / "build_pl_dashboard.py")
builder = importlib.util.module_from_spec(spec)
spec.loader.exec_module(builder)


def build_aux_maps(wb):
    client_map = {}
    for row in wb["Auxiliar_Cliente"].iter_rows(min_row=2, values_only=True):
        project = builder.clean(row[0], "")
        client = builder.clean(row[1], "")
        if project and client:
            client_map[project] = client

    razao_ws = wb["Auxiliar_Razao"]
    headers = [builder.norm(cell.value) for cell in next(razao_ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: index for index, header in enumerate(headers) if header}
    launch_col = idx.get("LANCAMENTO", idx.get("LANÇAMENTO", idx.get("CATEGORIA", 0)))
    category_col = idx.get("CATEGORIA", launch_col)
    account_col = idx.get("CONTA DO DRE", 3)
    percent_col = idx.get("ALIQUOTA", idx.get("ALÍQUOTA"))
    dre_map = defaultdict(list)
    for row in razao_ws.iter_rows(min_row=2, values_only=True):
        launch = builder.clean(row[launch_col] if launch_col < len(row) else None, "")
        if not launch:
            continue
        category = builder.clean(row[category_col] if category_col < len(row) else None, launch)
        account = builder.clean(row[account_col] if account_col < len(row) else None, "")
        percent_raw = row[percent_col] if percent_col is not None and percent_col < len(row) else None
        percent = builder.safe_float(percent_raw) if percent_raw not in (None, "") else builder.default_percent_for_category(category)
        dre_map[builder.norm(launch)].append(
            {
                "launch": launch,
                "category": category,
                "account": account,
                "percent": percent,
            }
        )
    return client_map, dre_map


def ledger_index():
    rows = json.loads(LEDGER.read_text(encoding="utf-8"))
    by_source_row = defaultdict(lambda: {
        "entries": 0,
        "sum_absorbed": 0.0,
        "gross": 0.0,
        "accounts": Counter(),
        "categories": Counter(),
        "clients": Counter(),
        "projects": Counter(),
        "periods": Counter(),
    })
    gross_by_client = defaultdict(float)
    gross_by_project = defaultdict(float)
    gross_by_period = defaultdict(float)
    for row in rows:
        if row.get("source") != "Faturamento_Pagamento":
            continue
        source_row = row.get("sourceRow")
        if not source_row:
            continue
        item = by_source_row[source_row]
        value = row.get("value") or 0.0
        item["entries"] += 1
        item["sum_absorbed"] += value
        item["accounts"][row.get("account") or ""] += 1
        item["categories"][row.get("category") or ""] += 1
        item["clients"][row.get("client") or ""] += 1
        item["projects"][row.get("project") or ""] += 1
        item["periods"][row.get("period") or ""] += 1
        if builder.norm(row.get("account")) == builder.norm("01. Receita Bruta de Vendas"):
            item["gross"] += value
            gross_by_client[row.get("client") or ""] += value
            gross_by_project[row.get("project") or ""] += value
            gross_by_period[row.get("period") or ""] += value
    return rows, by_source_row, gross_by_client, gross_by_project, gross_by_period


def common_value(counter):
    return counter.most_common(1)[0][0] if counter else ""


def joined(counter, limit=6):
    values = [name for name, _ in counter.most_common(limit) if name]
    extra = max(0, len(counter) - len(values))
    return ", ".join(values) + (f" (+{extra})" if extra else "")


def apply_sheet_style(ws, freeze="A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="F3F6FB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="42526E")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        1: 14,
        2: 12,
        3: 18,
        4: 18,
        5: 22,
        6: 22,
        7: 18,
        8: 16,
        9: 18,
        10: 16,
        11: 18,
        12: 22,
        13: 26,
        14: 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def write_table(wb_out, title, headers, rows, freeze="A2"):
    ws = wb_out.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    apply_sheet_style(ws, freeze)
    return ws


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    client_map, dre_map = build_aux_maps(wb)
    _, by_source_row, gross_by_client, gross_by_project, gross_by_period = ledger_index()

    ws = wb["Faturamento_Pagamento"]
    fp_idx = builder.header_index(ws)

    audited = []
    no_period = []
    no_depara = []
    not_in_ledger = []
    status_counts = Counter()
    launch_counts = Counter()
    totals = Counter()

    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue

        dt = builder.get_by_header(row, fp_idx, "DATA")
        period = builder.period_from_parts(
            builder.get_by_header(row, fp_idx, "Ano Competência", "Ano Competencia"),
            builder.get_by_header(row, fp_idx, "Mês Competência", "Mes Competencia"),
        )
        if not period and isinstance(dt, datetime):
            period = f"{dt.year}-{dt.month:02d}"

        launch = builder.clean(builder.get_by_header(row, fp_idx, "LANÇAMENTO", "LANCAMENTO"), "")
        project = builder.clean(builder.get_by_header(row, fp_idx, "OPERACAO", "OPERAÇÃO"), "")
        client = client_map.get(project, project)
        hub = builder.canonical_location(builder.clean(builder.get_by_header(row, fp_idx, "HUB"), ""))
        city = builder.clean(builder.get_by_header(row, fp_idx, "CIDADE"), "")
        vehicle_type = builder.clean(builder.get_by_header(row, fp_idx, "TIPO"), "")
        value = builder.safe_float(builder.get_by_header(row, fp_idx, "VALOR"))
        invoice = builder.clean(builder.get_by_header(row, fp_idx, "INVOICE", "NF"), "")
        mapped = builder.norm(launch) in dre_map
        ledger_item = by_source_row.get(excel_row)

        totals["linhas_fisicas"] += 1
        totals["valor_original"] += value
        launch_counts[launch or "(vazio)"] += 1

        if not period:
            status = "Não considerada - sem período"
            totals["sem_periodo"] += 1
        elif abs(value) <= 0.0001:
            status = "Operacional sem impacto financeiro - valor zero"
            totals["valor_zero"] += 1
        elif ledger_item:
            status = "Considerada no P&L"
            totals["consideradas_financeiro"] += 1
        else:
            status = "Atenção - não localizada no ledger"
            totals["nao_localizadas_ledger"] += 1

        if period:
            totals["com_periodo"] += 1
        if mapped:
            totals["com_depara"] += 1
        else:
            totals["sem_depara"] += 1
        if abs(value) > 0.0001:
            totals["valor_diferente_zero"] += 1

        status_counts[status] += 1

        entries = ledger_item["entries"] if ledger_item else 0
        absorbed = ledger_item["sum_absorbed"] if ledger_item else 0.0
        gross = ledger_item["gross"] if ledger_item else 0.0
        accounts = joined(ledger_item["accounts"]) if ledger_item else ""
        categories = joined(ledger_item["categories"]) if ledger_item else ""

        audit_row = [
            excel_row,
            status,
            period or "",
            launch,
            client,
            project,
            hub,
            city,
            vehicle_type,
            invoice,
            value,
            "Sim" if mapped else "Não",
            entries,
            absorbed,
            gross,
            accounts,
            categories,
        ]
        audited.append(audit_row)
        if not period:
            no_period.append(audit_row)
        if not mapped:
            no_depara.append(audit_row)
        if status == "Atenção - não localizada no ledger":
            not_in_ledger.append(audit_row)

    wb_out = Workbook()
    summary = wb_out.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Valor"])
    summary_rows = [
        ("Linhas físicas na Faturamento_Pagamento", totals["linhas_fisicas"]),
        ("Linhas com período identificado", totals["com_periodo"]),
        ("Linhas sem período e ignoradas pelo sistema", totals["sem_periodo"]),
        ("Linhas com valor diferente de zero", totals["valor_diferente_zero"]),
        ("Linhas com valor zero", totals["valor_zero"]),
        ("Linhas consideradas no financeiro do P&L", totals["consideradas_financeiro"]),
        ("Linhas não localizadas no ledger", totals["nao_localizadas_ledger"]),
        ("Linhas com de/para no Auxiliar_Razao", totals["com_depara"]),
        ("Linhas sem de/para no Auxiliar_Razao", totals["sem_depara"]),
        ("Valor original total da aba", totals["valor_original"]),
        ("Receita Bruta considerada no ledger", sum(gross_by_period.values())),
    ]
    for row in summary_rows:
        summary.append(row)
    summary.append([])
    summary.append(["Status", "Linhas"])
    for status, count in status_counts.most_common():
        summary.append([status, count])
    summary.append([])
    summary.append(["Top lançamentos por volume", "Linhas"])
    for launch, count in launch_counts.most_common(20):
        summary.append([launch, count])
    apply_sheet_style(summary)
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 18

    headers = [
        "Linha Excel",
        "Status",
        "Período",
        "Lançamento",
        "Cliente",
        "Projeto",
        "Hub",
        "Cidade",
        "Tipo",
        "NF",
        "Valor Original",
        "Tem De/Para",
        "Qtd. Ledger",
        "Valor Absorvido",
        "Receita Bruta",
        "Contas P&L",
        "Categorias P&L",
    ]
    write_table(wb_out, "Linhas auditadas", headers, audited)
    write_table(wb_out, "Sem depara", headers, no_depara)
    write_table(wb_out, "Sem periodo", headers, no_period)
    write_table(wb_out, "Nao localizadas ledger", headers, not_in_ledger)

    write_table(
        wb_out,
        "Receita por cliente",
        ["Cliente", "Receita Bruta Considerada"],
        sorted(gross_by_client.items(), key=lambda item: item[1], reverse=True),
    )
    write_table(
        wb_out,
        "Receita por projeto",
        ["Projeto", "Receita Bruta Considerada"],
        sorted(gross_by_project.items(), key=lambda item: item[1], reverse=True),
    )
    write_table(
        wb_out,
        "Receita por periodo",
        ["Período", "Receita Bruta Considerada"],
        sorted(gross_by_period.items()),
    )

    for ws_out in wb_out.worksheets:
        for row in ws_out.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

    wb_out.save(OUT)
    print(OUT)
    print("linhas_fisicas", totals["linhas_fisicas"])
    print("consideradas_financeiro", totals["consideradas_financeiro"])
    print("sem_periodo", totals["sem_periodo"])
    print("valor_zero", totals["valor_zero"])
    print("sem_depara", totals["sem_depara"])
    print("nao_localizadas_ledger", totals["nao_localizadas_ledger"])


if __name__ == "__main__":
    main()
