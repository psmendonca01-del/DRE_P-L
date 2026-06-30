import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd


DEFAULT_FOLDER = Path(
    r"C:\Users\PauloMendonça\OneDrive - Redefrete\Área de Trabalho\Balanço\DashBoard_P&L\Meli\012026"
)


def strip_accents(value):
    text = str(value or "")
    return "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")


def norm_text(value):
    return re.sub(r"\s+", " ", strip_accents(value).strip()).upper()


def norm_col(value):
    return re.sub(r"[^A-Z0-9]+", "", norm_text(value))


def clean_plate(value):
    text = re.sub(r"[^A-Z0-9-]", "", norm_text(value))
    if text.startswith("SDD-"):
        text = text[4:]
    return text


def clean_route_id(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"[^0-9A-Za-z]+", "", text)


def as_number(value):
    if pd.isna(value) or str(value).strip() == "":
        return 0.0
    text = str(value).strip()
    if text.upper() in {"#N/A", "NAN", "NONE"}:
        return 0.0
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value):
    if pd.isna(value) or str(value).strip() == "":
        return ""
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return text


def vehicle_from_description(description):
    text = norm_text(description)
    if text.startswith("VEICULO DE PASSEIO") or " PASSEIO " in f" {text} ":
        return "Passeio"
    if text.startswith("UTILITARIOS") or text.startswith("UTILITARIO"):
        return "Utilitario"
    if text.startswith("VUC") or " VUC " in f" {text} ":
        return "Vuc"
    if text.startswith("VAN") or " VAN " in f" {text} ":
        return "Van"
    return ""


def svc_from_description(description):
    match = re.search(r"SVC:\s*([A-Za-z0-9_\-]+)", str(description or ""), flags=re.I)
    return match.group(1).upper() if match else ""


def record_type(description):
    text = norm_text(description)
    if "VISITED ADDRESSES" in text:
        return "Complemento - Visited Addresses"
    if "LOST PACKAGES" in text:
        return "Deducao - Lost Packages"
    if "PNR PACKAGES" in text:
        return "Deducao - Pnr Packages"
    if "VEHICLE DAILY NOT VISITED" in text:
        return "Deducao - Vehicle Daily Not Visited"
    if "SVC:" in text:
        return "Rota"
    return "Nao classificado"


def is_subtotal_row(description):
    text = norm_col(description)
    return "SUBTOTAL" in text


def is_deduction(kind):
    return str(kind).startswith("Deducao")


def period_from_file(path):
    match = re.search(r"(20\d{2})(\d{2})Q0?(\d+)", path.stem, flags=re.I)
    if match:
        return f"{match.group(1)}-{match.group(2)}", f"Q{match.group(3)}"
    match = re.search(r"(\d{2})S(\d{2})(20\d{2})", path.stem, flags=re.I)
    if match:
        return f"{match.group(3)}-{match.group(2)}", f"S{match.group(1)}"
    return "", ""


def find_column(columns, *candidates):
    normalized = {norm_col(col): col for col in columns}
    for candidate in candidates:
        key = norm_col(candidate)
        if key in normalized:
            return normalized[key]
    for col in columns:
        col_norm = norm_col(col)
        if any(norm_col(candidate) in col_norm for candidate in candidates):
            return col
    return None


def read_meli_csv(path):
    period, bucket = period_from_file(path)
    preview = pd.read_csv(path, sep=";", encoding="utf-8", dtype=str, header=None, nrows=20)
    header_index = 0
    for idx, row in preview.iterrows():
        values = [norm_col(value) for value in row.tolist()]
        if "DESCRICAO" in values and "IDDAROTA" in values:
            header_index = idx
            break
    df = pd.read_csv(path, sep=";", encoding="utf-8", dtype=str, header=header_index)
    df = df.dropna(how="all").copy()
    cols = list(df.columns)
    desc_col = find_column(cols, "Descrição", "Descricao")
    route_col = find_column(cols, "ID da rota")
    start_col = find_column(cols, "Data de início", "Data inicio")
    end_col = find_column(cols, "Data de término", "Data termino")
    plate_col = find_column(cols, "Placa")
    driver_col = find_column(cols, "Motorista")
    qty_col = find_column(cols, "Quantidade")
    unit_col = find_column(cols, "Preço unitário", "Preco unitario")
    total_col = find_column(cols, "Total")
    iss_col = find_column(cols, "Total com ISS")
    icms_col = find_column(cols, "Total com ICMS")
    tax_col = find_column(cols, "Total com Tax")
    net_col = cols[-1] if len(cols) > 13 and str(cols[-1]).startswith("Unnamed") else tax_col
    line_col = cols[0] if norm_col(cols[0]) not in {"DESCRICAO"} else None

    base = pd.DataFrame(
        {
            "Arquivo CSV": path.name,
            "Periodo Arquivo": period,
            "Bloco Arquivo": bucket,
            "Linha Origem": df[line_col] if line_col else "",
            "Descricao Original": df[desc_col],
            "ID da Rota": df[route_col].map(clean_route_id),
            "Data Inicio": df[start_col].map(parse_date),
            "Data Termino": df[end_col].map(parse_date),
            "Placa": df[plate_col],
            "Placa Normalizada": df[plate_col].map(clean_plate),
            "Motorista": df[driver_col],
            "Motorista Normalizado": df[driver_col].map(norm_text),
            "Quantidade CSV": df[qty_col].map(as_number),
            "Preco Unitario": df[unit_col].map(as_number),
            "Total CSV": df[total_col].map(as_number),
            "Total ISS": df[iss_col].map(as_number),
            "Total ICMS": df[icms_col].map(as_number),
            "Total Tax": df[tax_col].map(as_number),
            "Valor Base": df[net_col].map(as_number),
        }
    )
    base = base[~base["Descricao Original"].map(is_subtotal_row)].copy()
    base["Tipo Registro"] = base["Descricao Original"].map(record_type)
    base["Eh Deducao"] = base["Tipo Registro"].map(is_deduction)
    base["Tipo Veiculo Extraido"] = base["Descricao Original"].map(vehicle_from_description)
    base["SVC Extraido"] = base["Descricao Original"].map(svc_from_description)
    base["Valor Sinalizado"] = base.apply(
        lambda row: -abs(row["Total CSV"]) if row["Eh Deducao"] else row["Total CSV"],
        axis=1,
    )
    base = base[base["ID da Rota"].ne("") | base["Tipo Registro"].ne("Nao classificado")].copy()
    return base


def enrich_csv_rows(base):
    routes = base[base["Tipo Registro"].eq("Rota")].copy()
    by_id = {}
    by_plate = {}
    by_driver = {}
    for _, row in routes.iterrows():
        if row["ID da Rota"] and row["ID da Rota"] not in by_id:
            by_id[row["ID da Rota"]] = row
        if row["Placa Normalizada"] and row["Placa Normalizada"] not in by_plate:
            by_plate[row["Placa Normalizada"]] = row
        if row["Motorista Normalizado"] and row["Motorista Normalizado"] not in by_driver:
            by_driver[row["Motorista Normalizado"]] = row

    enriched = []
    for _, row in base.iterrows():
        if row["Tipo Registro"] == "Rota":
            match = row
            criteria = "Proprio registro"
            found = "Rota principal"
        elif row["ID da Rota"] and row["ID da Rota"] in by_id:
            match = by_id[row["ID da Rota"]]
            criteria = "ID da rota"
            found = "Sim"
        elif row["Placa Normalizada"] and row["Placa Normalizada"] in by_plate:
            match = by_plate[row["Placa Normalizada"]]
            criteria = "Placa normalizada"
            found = "Sim"
        elif row["Motorista Normalizado"] and row["Motorista Normalizado"] in by_driver:
            match = by_driver[row["Motorista Normalizado"]]
            criteria = "Motorista"
            found = "Sim"
        else:
            match = None
            criteria = "Sem ID/placa/motorista correspondente"
            found = "Nao - orfao"

        if match is None:
            enriched.append(
                {
                    "Vinculo Encontrado": found,
                    "Criterio Vinculo": criteria,
                    "ID Rota Vinculada": "",
                    "SVC Final": "",
                    "Tipo Veiculo Final": "",
                    "Placa Rota Final": row["Placa Normalizada"],
                    "Motorista Rota Final": row["Motorista"],
                }
            )
        else:
            enriched.append(
                {
                    "Vinculo Encontrado": found,
                    "Criterio Vinculo": criteria,
                    "ID Rota Vinculada": match["ID da Rota"],
                    "SVC Final": match["SVC Extraido"],
                    "Tipo Veiculo Final": match["Tipo Veiculo Extraido"],
                    "Placa Rota Final": match["Placa Normalizada"],
                    "Motorista Rota Final": match["Motorista"],
                }
            )
    return pd.concat([base.reset_index(drop=True), pd.DataFrame(enriched)], axis=1)


def fill_missing_vehicle_and_svc(df):
    result = df.copy()

    def is_blank(value):
        return pd.isna(value) or str(value).strip() == ""

    def lookup_from(source, key_col, value_col):
        if key_col not in source or value_col not in source:
            return {}
        subset = source[source[key_col].fillna("").astype(str).str.strip().ne("")]
        subset = subset[subset[value_col].fillna("").astype(str).str.strip().ne("")]
        return subset.groupby(key_col)[value_col].agg(first_non_blank).to_dict()

    id_svc = lookup_from(result, "ID da Rota", "SVC Extraido")
    id_vehicle = lookup_from(result, "ID da Rota", "Tipo Veiculo Extraido")
    plate_svc = lookup_from(result, "Placa Normalizada", "SVC Extraido")
    plate_vehicle = lookup_from(result, "Placa Normalizada", "Tipo Veiculo Extraido")

    if "ID Rota Vinculada" in result:
        id_svc.update({k: v for k, v in lookup_from(result, "ID Rota Vinculada", "SVC Final").items() if v})
        id_vehicle.update({k: v for k, v in lookup_from(result, "ID Rota Vinculada", "Tipo Veiculo Final").items() if v})
    if "SVC Carregamento" in result:
        id_svc.update({k: v for k, v in lookup_from(result, "ID Rota Vinculada", "SVC Carregamento").items() if v})
    if "Tipo Veiculo Carregamento" in result:
        id_vehicle.update({k: v for k, v in lookup_from(result, "ID Rota Vinculada", "Tipo Veiculo Carregamento").items() if v})

    for idx, row in result.iterrows():
        route_id = str(row.get("ID Rota Vinculada") or row.get("ID da Rota") or "").strip()
        plate = str(row.get("Placa Normalizada") or row.get("Placa Rota Final") or "").strip()

        if "SVC Final" in result and is_blank(row.get("SVC Final")):
            result.at[idx, "SVC Final"] = id_svc.get(route_id) or plate_svc.get(plate) or ""
        if "Tipo Veiculo Final" in result and is_blank(row.get("Tipo Veiculo Final")):
            result.at[idx, "Tipo Veiculo Final"] = id_vehicle.get(route_id) or plate_vehicle.get(plate) or ""

    return result


def read_table_sheet(path, sheet_name, header_row):
    period, bucket = period_from_file(path)
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_row - 1, dtype=str, engine="calamine")
    except Exception:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            for _ in range(header_row - 1):
                next(rows, None)
            header = [str(value).strip() if value is not None else "" for value in next(rows)]
            data = []
            for row in rows:
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                values = list(row[: len(header)])
                if len(values) < len(header):
                    values.extend([""] * (len(header) - len(values)))
                data.append(values)
            wb.close()
            df = pd.DataFrame(data, columns=header, dtype=str)
        except Exception:
            return pd.DataFrame()
    df = df.dropna(how="all").copy()
    df.columns = [str(col).strip() for col in df.columns]
    df["Arquivo XLSX"] = path.name
    df["Periodo Arquivo"] = period
    df["Bloco Arquivo"] = bucket
    return df


def sheet_names(path):
    try:
        xl = pd.ExcelFile(path, engine="calamine")
        return list(xl.sheet_names)
    except Exception:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            names = list(wb.sheetnames)
            wb.close()
            return names
        except Exception:
            return []


def svc_name_from_text(value):
    text = str(value or "").strip()
    if text.lower() in {"", "nan", "none"}:
        return ""
    if " - " in text:
        return text.split(" - ", 1)[1].strip()
    return text


def read_svc_mapping(folder):
    rows = []
    source_files = [
        path
        for path in sorted(folder.glob("*.xlsx"))
        if not path.name.startswith("~$") and not path.name.startswith("Base_Meli_Consolidada_")
    ]
    for path in source_files:
        names = sheet_names(path)
        sheet = next((name for name in names if norm_text(name).replace(" ", "_") == "DE_>_PARA"), None)
        if not sheet:
            continue
        df = read_table_sheet(path, sheet, 1)
        code_col = find_column(df.columns, "SVC Code")
        name_col = find_column(df.columns, "SVC")
        system_col = find_column(df.columns, "SVC Sistema")
        if not code_col:
            continue
        for _, row in df.iterrows():
            code = str(row.get(code_col) or "").strip().upper()
            if not code:
                continue
            name = svc_name_from_text(row.get(name_col)) if name_col else ""
            if not name and system_col:
                name = svc_name_from_text(row.get(system_col))
            if name:
                rows.append({"SVC Final": code, "HUB Nome": name})
    if not rows:
        return pd.DataFrame(columns=["SVC Final", "HUB Nome"])
    return pd.DataFrame(rows).drop_duplicates("SVC Final")


def read_operational_files(folder):
    rotas_frames = []
    carregamento_frames = []
    source_files = [
        path
        for path in sorted(folder.glob("*.xlsx"))
        if not path.name.startswith("~$") and not path.name.startswith("Base_Meli_Consolidada_")
    ]
    for path in source_files:
        names = sheet_names(path)
        if not names:
            continue
        sheet_map = {norm_text(sheet).replace(" ", "_"): sheet for sheet in names}
        rota_sheet = sheet_map.get("ROTAS_SYS")
        carga_sheet = sheet_map.get("CARREGAMENTO")
        if rota_sheet:
            df = read_table_sheet(path, rota_sheet, 2)
            rota_id = find_column(df.columns, "Rota", "Rotas")
            loaded = find_column(df.columns, "Pacotes Carregados")
            delivered = find_column(df.columns, "Total Entregue")
            hub = find_column(df.columns, "Hub")
            if rota_id:
                out = pd.DataFrame(
                    {
                        "ID da Rota": df[rota_id].map(clean_route_id),
                        "Carregados Rotas_SYS": df[loaded].map(as_number) if loaded else 0,
                        "Entregues Rotas_SYS": df[delivered].map(as_number) if delivered else 0,
                        "Hub Rotas_SYS": df[hub].map(svc_name_from_text) if hub else "",
                        "Arquivo Rotas_SYS": path.name,
                    }
                )
                rotas_frames.append(out[out["ID da Rota"].ne("")])
        if carga_sheet:
            df = read_table_sheet(path, carga_sheet, 5)
            rota_id = find_column(df.columns, "ID ROTA")
            loaded = find_column(df.columns, "EMBARCADOS")
            delivered = find_column(df.columns, "ENTREGUES")
            svc = find_column(df.columns, "SVC")
            city = find_column(df.columns, "CIDADE")
            vehicle = find_column(df.columns, "TIPO")
            launch = find_column(df.columns, "LANCAMENTO", "LANÇAMENTO")
            date_col = find_column(df.columns, "DATA")
            if rota_id:
                out = pd.DataFrame(
                    {
                        "ID da Rota": df[rota_id].map(clean_route_id),
                        "Carregados Carregamento": df[loaded].map(as_number) if loaded else 0,
                        "Entregues Carregamento": df[delivered].map(as_number) if delivered else 0,
                        "SVC Carregamento": df[svc] if svc else "",
                        "Cidade Carregamento": df[city] if city else "",
                        "Tipo Veiculo Carregamento": df[vehicle] if vehicle else "",
                        "Lancamento Carregamento": df[launch] if launch else "",
                        "Data Carregamento": df[date_col].map(parse_date) if date_col else "",
                        "Arquivo Carregamento": path.name,
                    }
                )
                carregamento_frames.append(out[out["ID da Rota"].ne("")])
    rotas = pd.concat(rotas_frames, ignore_index=True) if rotas_frames else pd.DataFrame()
    carregamento = pd.concat(carregamento_frames, ignore_index=True) if carregamento_frames else pd.DataFrame()
    return rotas, carregamento


def aggregate_operational(rotas, carregamento):
    frames = []
    if not rotas.empty:
        frames.append(
            rotas.groupby("ID da Rota", as_index=False).agg(
                {
                    "Carregados Rotas_SYS": "sum",
                    "Entregues Rotas_SYS": "sum",
                    "Hub Rotas_SYS": lambda x: first_non_blank(x),
                    "Arquivo Rotas_SYS": lambda x: " | ".join(sorted(set(map(str, x)))),
                }
            )
        )
    if not carregamento.empty:
        frames.append(
            carregamento.groupby("ID da Rota", as_index=False).agg(
                {
                    "Carregados Carregamento": "sum",
                    "Entregues Carregamento": "sum",
                    "SVC Carregamento": lambda x: first_non_blank(x),
                    "Cidade Carregamento": lambda x: first_non_blank(x),
                    "Tipo Veiculo Carregamento": lambda x: first_non_blank(x),
                    "Arquivo Carregamento": lambda x: " | ".join(sorted(set(map(str, x)))),
                }
            )
        )
    if not frames:
        return pd.DataFrame()
    result = frames[0]
    for frame in frames[1:]:
        result = result.merge(frame, on="ID da Rota", how="outer")
    for col in ["Carregados Rotas_SYS", "Entregues Rotas_SYS", "Carregados Carregamento", "Entregues Carregamento"]:
        if col in result:
            result[col] = result[col].fillna(0)
    result["Carregados Final"] = result.get("Carregados Carregamento", 0)
    result["Entregues Final"] = result.get("Entregues Carregamento", 0)
    if "Carregados Rotas_SYS" in result:
        result["Carregados Final"] = result["Carregados Final"].where(result["Carregados Final"].ne(0), result["Carregados Rotas_SYS"])
    if "Entregues Rotas_SYS" in result:
        result["Entregues Final"] = result["Entregues Final"].where(result["Entregues Final"].ne(0), result["Entregues Rotas_SYS"])
    return result


def first_non_blank(values):
    for value in values:
        if pd.notna(value) and str(value).strip():
            return value
    return ""


def format_workbook(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B1F4D")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = 10
            for cell in list(col)[:300]:
                width = max(width, len(str(cell.value or "")) + 2)
            ws.column_dimensions[letter].width = min(width, 42)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
    wb.save(path)


def writable_output_path(path):
    if not path.exists():
        return path
    try:
        with path.open("a+b"):
            return path
    except PermissionError:
        return path.with_name(f"{path.stem}_novo{path.suffix}")


def build_copy_sheet(consolidated):
    hub = consolidated["HUB Nome"] if "HUB Nome" in consolidated else consolidated["SVC Final"]
    copy = pd.DataFrame(
        {
            "LANCAMENTO": consolidated["Tipo Registro"],
            "DATA": consolidated["Data Inicio"],
            "ID ROTA": consolidated["ID Rota Vinculada"],
            "NOME_MOTORISTA": consolidated["Motorista Rota Final"],
            "PLACA": consolidated["Placa Rota Final"],
            "TIPO": consolidated["Tipo Veiculo Final"],
            "OPERACAO": "MELI",
            "HUB": hub,
            "CIDADE": consolidated.get("Cidade Carregamento", ""),
            "EMBARCADOS": consolidated["Carregados Final"],
            "ENTREGUES": consolidated["Entregues Final"],
            "PERFORMANCE": consolidated["Performance Entrega"],
            "EVIDENCIADOS": "",
            "OBS": consolidated["Criterio Vinculo"].fillna("").astype(str)
            + " | "
            + consolidated["Descricao Original"].fillna("").astype(str),
            "VALOR": consolidated["Valor Sinalizado"],
        }
    )
    copy["DATA"] = pd.to_datetime(copy["DATA"], errors="coerce")
    return copy


def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FOLDER
    output = folder / f"Base_Meli_Consolidada_{folder.name}.xlsx"
    output = writable_output_path(output)

    csv_frames = [read_meli_csv(path) for path in sorted(folder.glob("*.csv"))]
    if not csv_frames:
        raise SystemExit(f"Nenhum CSV encontrado em {folder}")
    csv_base = fill_missing_vehicle_and_svc(enrich_csv_rows(pd.concat(csv_frames, ignore_index=True)))

    rotas_sys, carregamento = read_operational_files(folder)
    operational = aggregate_operational(rotas_sys, carregamento)
    consolidated = csv_base.merge(operational, left_on="ID Rota Vinculada", right_on="ID da Rota", how="left", suffixes=("", " Operacional"))
    consolidated = fill_missing_vehicle_and_svc(consolidated)
    svc_mapping = read_svc_mapping(folder)
    consolidated = consolidated.merge(svc_mapping, on="SVC Final", how="left")
    for fallback_col in ["SVC Carregamento", "Hub Rotas_SYS", "SVC Final"]:
        if fallback_col in consolidated:
            hub_text = consolidated["HUB Nome"].fillna("").astype(str).str.strip()
            hub_blank = hub_text.eq("") | hub_text.str.lower().isin(["nan", "none"])
            fallback = consolidated[fallback_col].fillna("").astype(str).str.strip()
            fallback = fallback.mask(fallback.str.lower().isin(["nan", "none"]), "")
            consolidated.loc[hub_blank & fallback.ne(""), "HUB Nome"] = fallback
    consolidated["Carregados Final"] = consolidated["Carregados Final"].fillna(0)
    consolidated["Entregues Final"] = consolidated["Entregues Final"].fillna(0)
    non_routes = ~consolidated["Tipo Registro"].eq("Rota")
    consolidated.loc[non_routes, ["Carregados Final", "Entregues Final"]] = 0
    consolidated["Performance Entrega"] = consolidated.apply(
        lambda row: row["Entregues Final"] / row["Carregados Final"] if row["Carregados Final"] else 0,
        axis=1,
    )
    route_level = (
        consolidated[consolidated["ID Rota Vinculada"].ne("")]
        .groupby("ID Rota Vinculada", as_index=False)
        .agg(
            {
                "Periodo Arquivo": lambda x: first_non_blank(x),
                "SVC Final": lambda x: first_non_blank(x),
                "Tipo Veiculo Final": lambda x: first_non_blank(x),
                "Placa Rota Final": lambda x: first_non_blank(x),
                "Motorista Rota Final": lambda x: first_non_blank(x),
                "Valor Sinalizado": "sum",
                "Carregados Final": "max",
                "Entregues Final": "max",
                "Arquivo CSV": lambda x: " | ".join(sorted(set(map(str, x)))),
                "Arquivo Carregamento": lambda x: " | ".join(sorted(set(str(v) for v in x if pd.notna(v) and str(v).strip()))),
                "Arquivo Rotas_SYS": lambda x: " | ".join(sorted(set(str(v) for v in x if pd.notna(v) and str(v).strip()))),
            }
        )
    )
    route_level["Performance Entrega"] = route_level.apply(
        lambda row: row["Entregues Final"] / row["Carregados Final"] if row["Carregados Final"] else 0,
        axis=1,
    )

    summary = consolidated.groupby("Tipo Registro", as_index=False).agg(
        Linhas=("Tipo Registro", "size"),
        Valor=("Valor Sinalizado", "sum"),
        Rotas=("ID Rota Vinculada", "nunique"),
        Carregados=("Carregados Final", "sum"),
        Entregues=("Entregues Final", "sum"),
    )
    summary["Performance"] = summary.apply(lambda row: row["Entregues"] / row["Carregados"] if row["Carregados"] else 0, axis=1)

    svc_summary = consolidated.groupby(["SVC Final", "Tipo Veiculo Final", "Tipo Registro"], dropna=False, as_index=False).agg(
        Linhas=("Tipo Registro", "size"),
        Valor=("Valor Sinalizado", "sum"),
        Rotas=("ID Rota Vinculada", "nunique"),
        Carregados=("Carregados Final", "sum"),
        Entregues=("Entregues Final", "sum"),
    )
    svc_summary["Performance"] = svc_summary.apply(lambda row: row["Entregues"] / row["Carregados"] if row["Carregados"] else 0, axis=1)

    orphans = consolidated[consolidated["Vinculo Encontrado"].eq("Nao - orfao")].copy()
    no_ops = consolidated[
        consolidated["Tipo Registro"].eq("Rota")
        & consolidated["ID Rota Vinculada"].ne("")
        & consolidated["Carregados Final"].eq(0)
        & consolidated["Entregues Final"].eq(0)
    ].copy()
    copy_sheet = build_copy_sheet(consolidated)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        copy_sheet.to_excel(writer, sheet_name="Para Copiar", index=False)
        summary.to_excel(writer, sheet_name="Resumo", index=False)
        svc_summary.to_excel(writer, sheet_name="Resumo SVC", index=False)
        route_level.to_excel(writer, sheet_name="Base por Rota", index=False)
        consolidated.to_excel(writer, sheet_name="Base Consolidada", index=False)
        csv_base.to_excel(writer, sheet_name="Base CSV Tratada", index=False)
        rotas_sys.to_excel(writer, sheet_name="Rotas_SYS", index=False)
        carregamento.to_excel(writer, sheet_name="Carregamento", index=False)
        operational.to_excel(writer, sheet_name="Operacional por Rota", index=False)
        orphans.to_excel(writer, sheet_name="Orfaos", index=False)
        no_ops.to_excel(writer, sheet_name="Sem Operacional", index=False)
        pd.DataFrame(
            [
                {"Regra": "Arquivos CSV", "Aplicacao": "Todos os CSVs da pasta foram tratados e unidos."},
                {"Regra": "Arquivos XLSX", "Aplicacao": "Abas ROTAS SYS e CARREGAMENTO foram lidas com cabeçalho detectado nas linhas 2 e 5."},
                {"Regra": "ID da rota", "Aplicacao": "CSV usa ID da rota; ROTAS SYS usa Rota; CARREGAMENTO usa ID ROTA."},
                {"Regra": "Carregados/Entregues", "Aplicacao": "Prioriza CARREGAMENTO; se estiver zerado, usa ROTAS SYS."},
                {"Regra": "Vínculo", "Aplicacao": "Complementos e deduções vinculam por ID, depois placa sem SDD, depois motorista."},
            ]
        ).to_excel(writer, sheet_name="Regras", index=False)

    format_workbook(output)
    print(output)
    xlsx_sources = [
        path
        for path in folder.glob("*.xlsx")
        if not path.name.startswith("~$") and not path.name.startswith("Base_Meli_Consolidada_")
    ]
    print(f"CSVs: {len(csv_frames)} | XLSX fonte: {len(xlsx_sources)}")
    print(f"Base consolidada: {len(consolidated)} linhas")
    print(f"Rotas_SYS: {len(rotas_sys)} linhas | Carregamento: {len(carregamento)} linhas | Operacional rotas: {len(operational)}")
    print(f"Base por rota: {len(route_level)} rotas | Orfaos: {len(orphans)} | Sem operacional: {len(no_ops)}")


if __name__ == "__main__":
    main()
